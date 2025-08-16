import sys
import os
import re
import webbrowser
import json
from datetime import datetime, timedelta, timezone
import math
from typing import Optional, List, Dict, Any
import io
import concurrent.futures
from threading import Lock
import babel.support
import babel.core
from babel.support import Translations

import requests
import pandas as pd
try:
    from PIL import Image
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Alignment
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QLineEdit, QComboBox, QSpinBox, QRadioButton, QGroupBox,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView, QCheckBox,
    QMessageBox, QAbstractItemView, QFileDialog, QDialog, QListWidget,
    QListWidgetItem, QTextEdit, QProgressDialog, QSizePolicy
)
from PySide6.QtCore import Qt, QThread, Signal, QTranslator, QLocale, QTimer
from PySide6.QtGui import QClipboard
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

import path_manager

# gettext 초기화
def setup_i18n(language='ko'):
    """국제화(i18n) 설정 - 완전히 Babel 기반"""
    global _
    
    # 한국어는 원본 텍스트를 그대로 사용
    if language == 'ko':
        def korean_translate(text):
            return text
        _ = korean_translate
        print(f"언어 설정 완료: {language}")
        return
    
    try:
        # 언어 파일 경로 설정
        locale_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'locales')
        
        # Babel Translations 객체 생성
        from babel.support import Translations
        catalog = Translations.load(locale_dir, locales=[language], domain='messages')
        
        # 전역 _ 함수 설정
        _ = catalog.gettext
        print(f"언어 설정 완료: {language}")
    except Exception as e:
        print(f"언어 설정 중 오류 발생: {e}")
        # 기본 fallback 함수
        def fallback_translate(text):
            return text
        _ = fallback_translate

# 초기 언어 설정은 main 함수에서 처리됩니다

# 두 번째 이미지와 동일한 UI 스타일
STYLESHEET = """
QMainWindow {
    background-color: #F5F5F5;
}
QWidget {
    background-color: #FFFFFF;
    color: #333333;
    font-family: 'Malgun Gothic', '맑은 고딕', sans-serif;
    font-size: 12px;
}
QLabel {
    background-color: transparent;
    color: #333333;
    font-weight: normal;
}
QLineEdit, QComboBox, QSpinBox {
    border: 1px solid #CCCCCC;
    border-radius: 2px;
    padding: 4px 8px;
    background-color: #FFFFFF;
    min-height: 22px;
    max-height: 22px;
}
QLineEdit:focus, QComboBox:focus, QSpinBox:focus {
    border: 1px solid #0078D4;
}
QComboBox::drop-down {
    border: none;
    width: 20px;
}
QComboBox::down-arrow {
    image: none;
    border: none;
}
QPushButton {
    background-color: #E1E1E1;
    color: #333333;
    border: 1px solid #ADADAD;
    border-radius: 2px;
    padding: 6px 12px;
    font-size: 12px;
}
QPushButton:hover {
    background-color: #D4D4D4;
}
QPushButton:pressed {
    background-color: #CCCCCC;
}
QPushButton:disabled {
    background-color: #F0F0F0;
    color: #999999;
    border: 1px solid #DDDDDD;
}
QGroupBox {
    border: none;
    padding-top: 10px;
    font-weight: normal;
}
QGroupBox::title {
    color: #333333;
    font-weight: normal;
}
QTableWidget {
    border: 1px solid #CCCCCC;
    gridline-color: #E0E0E0;
    background-color: #FFFFFF;
}
QHeaderView::section {
    background-color: #F0F0F0;
    padding: 4px;
    border: 1px solid #CCCCCC;
    font-weight: bold;
    font-size: 12px;
}
QRadioButton {
    spacing: 5px;
}
QRadioButton::indicator {
    width: 13px;
    height: 13px;
}
QCheckBox {
    spacing: 5px;
}
QCheckBox::indicator {
    width: 16px;
    height: 16px;
    border: 2px solid #CCCCCC;
    border-radius: 3px;
    background-color: #FFFFFF;
}
QCheckBox::indicator:checked {
    background-color: #0078D4;
    border: 2px solid #0078D4;
    image: url(data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTIiIGhlaWdodD0iOSIgdmlld0JveD0iMCAwIDEyIDkiIGZpbGw9Im5vbmUiIHhtbG5zPSJodHRwOi8vd3d3LnczLm9yZy8yMDAwL3N2ZyI+CjxwYXRoIGQ9Ik0xMC42IDEuNEw0LjIgNy44TDEuNCA1IiBzdHJva2U9IndoaXRlIiBzdHJva2Utd2lkdGg9IjIiIHN0cm9rZS1saW5lY2FwPSJyb3VuZCIgc3Ryb2tlLWxpbmVqb2luPSJyb3VuZCIvPgo8L3N2Zz4K);
    background-repeat: no-repeat;
    background-position: center;
}
QCheckBox::indicator:hover {
    border: 2px solid #0078D4;
}
/* 현대적인 스크롤바 스타일 */
QScrollBar:vertical {
    background-color: #F5F5F5;
    width: 12px;
    border-radius: 6px;
    margin: 0px;
}
QScrollBar::handle:vertical {
    background-color: #D0D0D0;
    border-radius: 6px;
    min-height: 30px;
    margin: 2px;
}
QScrollBar::handle:vertical:hover {
    background-color: #B0B0B0;
}
QScrollBar::handle:vertical:pressed {
    background-color: #909090;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0px;
}
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
    background: transparent;
}
QScrollBar:horizontal {
    background-color: #F5F5F5;
    height: 12px;
    border-radius: 6px;
    margin: 0px;
}
QScrollBar::handle:horizontal {
    background-color: #D0D0D0;
    border-radius: 6px;
    min-width: 30px;
    margin: 2px;
}
QScrollBar::handle:horizontal:hover {
    background-color: #B0B0B0;
}
QScrollBar::handle:horizontal:pressed {
    background-color: #909090;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    width: 0px;
}
QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
    background: transparent;
}
"""

class Worker(QThread):
    finished = Signal(dict)
    error = Signal(str)
    progress = Signal(str)

    def __init__(self, search_keyword, upload_period, min_views, sort_by, video_duration, max_subs, api_key, translator_func, similar_video_url=None, analyze_contribution=False):
        super().__init__()
        self.api_key = api_key
        self.search_keyword = search_keyword
        self._ = translator_func # 이 줄이 추가되었습니다.
        self.upload_period = upload_period
        self.min_views = min_views
        self.sort_by = sort_by
        self.video_duration = video_duration
        self.max_subs = max_subs
        self.similar_video_url = similar_video_url  # 유사 영상 검색용 URL
        self.analyze_contribution = analyze_contribution  # 채널 기여도 분석 여부
        self.youtube = None
        self._channel_cache = {}  # 채널 정보 캐시
        self._video_cache = {}    # 비디오 정보 캐시
        self._ = translator_func

    def _extract_video_id_from_url(self, url):
        """URL에서 비디오 ID를 추출 (쇼츠 URL 지원)"""
        import re
        
        if not url: 
            return None
            
        patterns = [
            # 쇼츠 URL을 가장 먼저 (최우선 처리)
            r'(?:https?://)?(?:www\.)?youtube\.com/shorts/([a-zA-Z0-9_-]{11})',
            r'(?:https?://)?(?:m\.)?youtube\.com/shorts/([a-zA-Z0-9_-]{11})',
            # 일반 URL들
            r'(?:https?://)?(?:www\.)?youtube\.com/watch\?.*v=([a-zA-Z0-9_-]{11})',
            r'(?:https?://)?youtu\.be/([a-zA-Z0-9_-]{11})',
            r'(?:https?://)?(?:www\.)?youtube\.com/embed/([a-zA-Z0-9_-]{11})',
            r'(?:https?://)?m\.youtube\.com/watch\?.*v=([a-zA-Z0-9_-]{11})'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        
        return None

    def _execute_robust_fallback_search(self):
        """견고한 순차적 폴백 검색 알고리즘 - 절대 멈추지 않는 검색"""
        try:
            # API 클라이언트 초기화
            self.youtube = build('youtube', 'v3', developerKey=self.api_key)
            
            # A. 사전 준비
            search_results_ids = []
            search_params = None
            page_count = 0
            
            self.progress.emit(self._("원본 영상 분석 중..."))
            
            # 원본 영상 메타데이터 확보
            video_id = self._extract_video_id_from_url(self.similar_video_url)
            if not video_id:
                raise ValueError(self._("올바른 유튜브 영상 URL을 입력해주세요."))
            
            response = self.youtube.videos().list(
                part='snippet,topicDetails,contentDetails',
                id=video_id
            ).execute()
            
            if not response.get('items'):
                raise ValueError(self._("해당 영상을 찾을 수 없습니다."))
            
            video_info = response['items'][0]
            snippet = video_info['snippet']
            
            # 메타데이터 정리
            title = snippet.get('title', '')
            category_id = snippet.get('categoryId', '')
            channel_id = snippet.get('channelId', '')
            default_language = snippet.get('defaultAudioLanguage', 'ko')
            topic_ids = video_info.get('topicDetails', {}).get('topicIds', [])
            
            # 언어-지역 매핑
            language_to_region = {
                'ko': 'KR', 'en': 'US', 'ja': 'JP', 'zh': 'CN', 'es': 'ES',
                'fr': 'FR', 'de': 'DE', 'ru': 'RU', 'pt': 'BR', 'it': 'IT'
            }
            region_code = language_to_region.get(default_language[:2], 'KR')
            
            # 사용자 필터 준비 (핵심!)
            search_time = datetime.now()
            published_after = self._get_published_after_date(self.upload_period, search_time)
            
            # B. 견고한 순차적 폴백 검색 실행
            self.progress.emit(self._("1차 시도: 정밀 타겟팅 검색..."))
            
            # 1차 시도: 정밀 타겟팅 (title + videoCategoryId + 모든 필터)
            if not search_results_ids:
                try:
                    params = {
                        'q': title,
                        'part': 'snippet',
                        'type': 'video',
                        'order': 'relevance',
                        'maxResults': 50,
                        'videoDuration': self.video_duration,
                        'relevanceLanguage': default_language[:2] if default_language else None,
                        'regionCode': region_code
                    }
                    if category_id:
                        params['videoCategoryId'] = category_id
                    if published_after:
                        params['publishedAfter'] = published_after.isoformat() + 'Z'
                    
                    # None 값 제거
                    params = {k: v for k, v in params.items() if v is not None}
                    
                    response = self.youtube.search().list(**params).execute()
                    search_results_ids = [item['id']['videoId'] for item in response.get('items', [])]
                    
                    if search_results_ids:
                        search_params = params  # 성공한 파라미터 저장
                        self.progress.emit(self._("1차 성공! {}개 결과 발견").format(len(search_results_ids)))
                except Exception as e:
                    self.progress.emit(self._("1차 시도 실패, 2차 시도 진행..."))
            
            # 2차 시도: 주제 확장 (topicId + 모든 필터)
            if not search_results_ids and topic_ids:
                try:
                    self.progress.emit(self._("2차 시도: 주제 확장 검색..."))
                    params = {
                        'topicId': topic_ids[0],
                        'part': 'snippet',
                        'type': 'video',
                        'order': 'relevance',
                        'maxResults': 50,
                        'videoDuration': self.video_duration,
                        'relevanceLanguage': default_language[:2] if default_language else None,
                        'regionCode': region_code
                    }
                    if published_after:
                        params['publishedAfter'] = published_after.isoformat() + 'Z'
                    
                    # None 값 제거
                    params = {k: v for k, v in params.items() if v is not None}
                    
                    response = self.youtube.search().list(**params).execute()
                    search_results_ids = [item['id']['videoId'] for item in response.get('items', [])]
                    
                    if search_results_ids:
                        search_params = params  # 성공한 파라미터 저장
                        self.progress.emit(self._("2차 성공! {}개 결과 발견").format(len(search_results_ids)))
                except Exception as e:
                    self.progress.emit(self._("2차 시도 실패, 3차 시도 진행..."))
            
            # 3차 시도: 채널 우선 (channelId + 모든 필터)
            if not search_results_ids:
                try:
                    self.progress.emit(self._("3차 시도: 채널 우선 검색..."))
                    params = {
                        'channelId': channel_id,
                        'part': 'snippet',
                        'type': 'video',
                        'order': 'relevance',
                        'maxResults': 50,
                        'videoDuration': self.video_duration
                    }
                    if published_after:
                        params['publishedAfter'] = published_after.isoformat() + 'Z'
                    
                    # None 값 제거
                    params = {k: v for k, v in params.items() if v is not None}
                    
                    response = self.youtube.search().list(**params).execute()
                    search_results_ids = [item['id']['videoId'] for item in response.get('items', [])]
                    
                    if search_results_ids:
                        search_params = params  # 성공한 파라미터 저장
                        self.progress.emit(self._("3차 성공! {}개 결과 발견").format(len(search_results_ids)))
                except Exception as e:
                    self.progress.emit(self._("3차 시도 실패, 4차 시도 진행..."))
            
            # 4차 시도: 제목만 (title + 언어/지역 필터만)
            if not search_results_ids:
                try:
                    self.progress.emit(self._("4차 시도: 제목 기본 검색..."))
                    params = {
                        'q': title,
                        'part': 'snippet',
                        'type': 'video',
                        'order': 'relevance',
                        'maxResults': 50,
                        'videoDuration': self.video_duration,
                        'relevanceLanguage': default_language[:2] if default_language else None,
                        'regionCode': region_code
                    }
                    
                    # None 값 제거
                    params = {k: v for k, v in params.items() if v is not None}
                    
                    response = self.youtube.search().list(**params).execute()
                    search_results_ids = [item['id']['videoId'] for item in response.get('items', [])]
                    
                    if search_results_ids:
                        search_params = params  # 성공한 파라미터 저장
                        self.progress.emit(self._("4차 성공! {}개 결과 발견").format(len(search_results_ids)))
                except Exception as e:
                    self.progress.emit(self._("4차 시도도 실패"))
            
            # C. 페이지네이션 구현 (성공한 검색이 있을 경우)
            if search_results_ids and search_params:
                next_page_token = response.get('nextPageToken')
                
                # 최대 3페이지 추가 수집
                while next_page_token and page_count < 3 and len(search_results_ids) < 200:
                    page_count += 1
                    self.progress.emit(self._("추가 페이지 {}/3 수집 중...").format(page_count))
                    
                    try:
                        search_params['pageToken'] = next_page_token
                        response = self.youtube.search().list(**search_params).execute()
                        
                        new_ids = [item['id']['videoId'] for item in response.get('items', [])]
                        for vid_id in new_ids:
                            if vid_id not in search_results_ids:
                                search_results_ids.append(vid_id)
                        
                        next_page_token = response.get('nextPageToken')
                        
                        if not next_page_token:
                            break
                            
                    except Exception as e:
                        break
            
            # 최종 결과 처리
            if not search_results_ids:
                self.finished.emit({'results': [], 'total_count': 0})
                return
            
            self.progress.emit(self._("총 {}개 영상 상세 정보 수집 중...").format(len(search_results_ids)))
            
            # 영상 상세 정보 수집
            all_video_details = self._get_video_details_parallel(search_results_ids)
            
            # 채널 정보 수집 (기여도 분석은 주문형으로 변경)
            channel_ids = list(set(v['snippet']['channelId'] for v in all_video_details))
            subscriber_counts = self._get_channel_stats_parallel(channel_ids)
            
            # 결과 가공 및 필터링
            self.progress.emit(self._("필터링 및 정렬 중..."))
            results = []
            
            for video in all_video_details:
                view_count = int(video['statistics'].get('viewCount', 0))
                like_count = int(video['statistics'].get('likeCount', 0))
                comment_count = int(video['statistics'].get('commentCount', 0))
                duration = video['contentDetails'].get('duration', 'PT0M0S')
                channel_id = video['snippet']['channelId']
                subscriber_count = subscriber_counts.get(channel_id, 0)
                license_info = video.get('status', {}).get('license', 'youtube')
                engagement_rate = like_count / max(view_count, 1)
                reaction_rate = comment_count / max(view_count, 1)
                
                if (view_count >= self.min_views and 
                    (self.max_subs == 0 or subscriber_count <= self.max_subs)):
                    
                    result_item = {
                        'video_id': video['id'],
                        'published_at': video['snippet']['publishedAt'].split('T')[0],
                        'published_at_full': video['snippet']['publishedAt'],
                        'view_count': view_count,
                        'like_count': like_count,
                        'comment_count': comment_count,
                        'duration': duration,
                        'duration_formatted': self._format_duration(duration),
                        'title': video['snippet']['title'],
                        'channel_title': video['snippet']['channelTitle'],
                        'channel_id': channel_id,
                        'subscriber_count': subscriber_count,
                        'engagement_rate': engagement_rate,
                        'reaction_rate': reaction_rate,
                        'license': license_info,
                        'thumbnail_url': video['snippet']['thumbnails'].get('medium', {}).get('url', 
                                       video['snippet']['thumbnails'].get('default', {}).get('url', ''))
                    }
                    
                    # 기여도 분석은 주문형으로 변경 - 기본 검색에서는 제거
                    
                    results.append(result_item)
            
            results = self._sort_results(results)
            
            self.progress.emit(self._("유사 영상 검색 완료!"))
            self.finished.emit({'results': results, 'total_count': len(results)})
                
        except Exception as e:
            self.error.emit(self._("견고한 검색 오류: ") + str(e))



















    def _calculate_channel_averages_optimized(self, channel_ids):
        """최적화된 채널별 평균 조회수 계산 (모든 채널 분석)"""
        channel_averages = {}
        total_channels = len(channel_ids)
        
        for idx, channel_id in enumerate(channel_ids):  # 모든 채널 분석 (제한 제거)
            try:
                # 진행상황 표시
                self.progress.emit(self._("채널 기여도 분석 중... ({}/{})").format(idx + 1, total_channels))
                
                # 채널의 최근 업로드 영상들 가져오기
                search_response = self.youtube.search().list(
                    channelId=channel_id,
                    part='snippet',
                    type='video',
                    order='date',
                    maxResults=20
                ).execute()
                
                recent_video_ids = [item['id']['videoId'] for item in search_response.get('items', [])]
                
                if recent_video_ids:
                    # 영상들의 조회수 가져오기
                    videos_response = self.youtube.videos().list(
                        part='statistics',
                        id=','.join(recent_video_ids)
                    ).execute()
                    
                    view_counts = []
                    for video in videos_response.get('items', []):
                        view_count = int(video['statistics'].get('viewCount', 0))
                        if view_count > 0:
                            view_counts.append(view_count)
                    
                    if view_counts:
                        channel_averages[channel_id] = sum(view_counts) // len(view_counts)
                        
            except Exception:
                continue
                
        return channel_averages


class ContributionWorker(QThread):
    """주문형 기여도 분석을 위한 별도 워커 클래스"""
    finished = Signal(int, dict)  # row, contribution_data
    error = Signal(str)
    
    def __init__(self, api_key, channel_id, video_id, view_count, row):
        super().__init__()
        self.api_key = api_key
        self.channel_id = channel_id
        self.video_id = video_id
        self.view_count = view_count
        self.row = row
        self.youtube = None
    
    def run(self):
        try:
            # API 클라이언트 초기화
            self.youtube = build('youtube', 'v3', developerKey=self.api_key)
            
            all_video_ids = []
            
            # 최근 업로드 영상들 가져오기 (최대 50개) - 채널의 '지금'을 보여주는 지표
            recent_response = self.youtube.search().list(
                channelId=self.channel_id,
                part='snippet',
                type='video',
                order='date',
                maxResults=50
            ).execute()
            
            recent_video_ids = [item['id']['videoId'] for item in recent_response.get('items', [])]
            all_video_ids.extend(recent_video_ids)
            
            if not all_video_ids:
                # 영상이 없으면 기본값 반환
                contribution_data = {
                    'contribution_score': 0,
                    'channel_avg_views': 0,
                    'error': 'No videos found'
                }
                self.finished.emit(self.row, contribution_data)
                return
            
            # 영상들의 조회수 가져오기 (최대 50개까지)
            video_chunks = [all_video_ids[i:i+50] for i in range(0, len(all_video_ids), 50)]
            view_counts = []
            
            for chunk in video_chunks:
                videos_response = self.youtube.videos().list(
                    part='statistics',
                    id=','.join(chunk)
                ).execute()
                
                for video in videos_response.get('items', []):
                    view_count = int(video['statistics'].get('viewCount', 0))
                    # 조회수가 0인 영상도 포함하여 더 정확한 평균 계산
                    view_counts.append(view_count)
            
            if not view_counts:
                # 조회수 정보가 없으면 기본값 반환
                contribution_data = {
                    'contribution_score': 0,
                    'channel_avg_views': 0,
                    'error': 'No view data available'
                }
                self.finished.emit(self.row, contribution_data)
                return
            
            # 채널 평균 조회수 계산 (더 정확한 평균)
            channel_avg_views = sum(view_counts) // len(view_counts)
            
            # 기여도 계산
            if channel_avg_views > 0:
                contribution_score = (self.view_count / channel_avg_views) * 100
            else:
                contribution_score = 0
            
            # 결과 데이터 구성
            contribution_data = {
                'contribution_score': contribution_score,
                'channel_avg_views': channel_avg_views
            }
            
            self.finished.emit(self.row, contribution_data)
            
        except Exception as e:
            # API 오류 처리
            from googleapiclient.errors import HttpError
            import json
            
            if isinstance(e, HttpError):
                try:
                    error_details = json.loads(e.content.decode('utf-8'))['error']
                    reason = error_details.get('errors', [{}])[0].get('reason', 'unknown')
                    
                    if reason == 'quotaExceeded':
                        self.error.emit("QUOTA_EXCEEDED")
                    elif reason == 'keyInvalid':
                        self.error.emit("KEY_INVALID")
                    else:
                        self.error.emit(f"API 오류: {error_details.get('message', str(e))}")
                except:
                    self.error.emit(f"API 오류: {str(e)}")
            else:
                self.error.emit(f"기여도 분석 오류: {str(e)}")


class SimilarVideoWorker(QThread):
    """유사 영상 검색 전용 워커 클래스"""
    progress = Signal(str)
    finished = Signal(dict)
    error = Signal(str)

    def __init__(self, video_url, upload_period, min_views, sort_by, video_duration, max_subs, api_key, translator_func):
        super().__init__()
        self.video_url = video_url
        self.upload_period = upload_period
        self.min_views = min_views
        self.sort_by = sort_by
        self.video_duration = video_duration
        self.max_subs = max_subs
        self.api_key = api_key
        self._ = translator_func
        self.youtube = None
        # [Task 3 Solution] Add a cancellation flag
        self.is_cancelled = False

    # [Task 3 Solution] Add a method to set the cancellation flag
    def cancel(self):
        """Signals the worker thread to stop its operation."""
        self.is_cancelled = True
        print("[INFO] Search cancellation requested by user.")

    def _extract_video_id_from_url(self, url):
        """URL에서 비디오 ID를 추출 (쇼츠 URL 지원)"""
        import re
        
        if not url: 
            return None
            
        patterns = [
            # 쇼츠 URL을 가장 먼저 (최우선 처리)
            r'(?:https?://)?(?:www\.)?youtube\.com/shorts/([a-zA-Z0-9_-]{11})',
            r'(?:https?://)?(?:m\.)?youtube\.com/shorts/([a-zA-Z0-9_-]{11})',
            # 일반 URL들
            r'(?:https?://)?(?:www\.)?youtube\.com/watch\?.*v=([a-zA-Z0-9_-]{11})',
            r'(?:https?://)?youtu\.be/([a-zA-Z0-9_-]{11})',
            r'(?:https?://)?(?:www\.)?youtube\.com/embed/([a-zA-Z0-9_-]{11})',
            r'(?:https?://)?m\.youtube\.com/watch\?.*v=([a-zA-Z0-9_-]{11})'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        
        return None

    def _search_with_retry(self, params):
        """네트워크 오류 방지를 위한 재시도 기능이 탑재된 안정적인 검색 API 호출 함수"""
        import time
        
        if self.is_cancelled: return []
        
        retry_delays = [2, 4]  # 2초, 4초 간격으로 재시도
        last_error = None
        
        for attempt in range(3):  # 최대 3번 시도
            try:
                # YouTube 클라이언트 재생성으로 연결 안정성 확보
                if attempt > 0:
                    self.youtube = build('youtube', 'v3', developerKey=self.api_key)
                
                response = self.youtube.search().list(**params).execute()
                items = response.get('items', [])
                return items
                
            except KeyboardInterrupt:
                # 사용자가 강제 종료한 경우
                print("    [중단] 사용자에 의한 검색 중단")
                return []
            except Exception as e:
                last_error = e
                error_msg = str(e)
                print(f"    [재시도 {attempt + 1}/3] 검색 API 호출 실패: {error_msg}")
                
                # 특정 오류는 재시도하지 않음 (잘못된 파라미터 등)
                if ("badRequest" in error_msg or "quotaExceeded" in error_msg or 
                    "keyInvalid" in error_msg or "invalid argument" in error_msg):
                    print(f"    [중단] 복구 불가능한 오류: {error_msg}")
                    return []
                
                if attempt < 2:  # 마지막 시도가 아니면 대기
                    time.sleep(retry_delays[attempt])
                    print(f"    {retry_delays[attempt]}초 후 재시도...")
        
        print(f"    [최종 실패] 3번 시도 후에도 실패: {str(last_error)}")
        return []

    def _videos_with_retry(self, video_ids):
        """네트워크 오류 방지를 위한 재시도 기능이 탑재된 안정적인 비디오 API 호출 함수"""
        import time
        
        if not video_ids or self.is_cancelled: return []
        
        retry_delays = [2, 4]  # 2초, 4초 간격으로 재시도
        last_error = None
        
        for attempt in range(3):  # 최대 3번 시도
            try:
                # YouTube 클라이언트 재생성으로 연결 안정성 확보
                if attempt > 0:
                    self.youtube = build('youtube', 'v3', developerKey=self.api_key)
                
                response = self.youtube.videos().list(
                    part='snippet,statistics,contentDetails,status',
                    id=','.join(video_ids)
                ).execute()
                
                items = response.get('items', [])
                if items:  # 성공적으로 데이터를 받았으면 즉시 반환
                    return items
                else:
                    print(f"    [경고] 비디오 API 응답이 비어있음 (시도 {attempt + 1}/3)")
                    
            except KeyboardInterrupt:
                # 사용자가 강제 종료한 경우
                print("    [중단] 사용자에 의한 검색 중단")
                return []
            except Exception as e:
                last_error = e
                error_msg = str(e)
                print(f"    [재시도 {attempt + 1}/3] 비디오 API 호출 실패: {error_msg}")
                
                # 특정 오류는 재시도하지 않음
                if "quotaExceeded" in error_msg or "keyInvalid" in error_msg:
                    print(f"    [중단] 복구 불가능한 오류: {error_msg}")
                    return []
                
                if attempt < 2:  # 마지막 시도가 아니면 대기
                    time.sleep(retry_delays[attempt])
                    print(f"    {retry_delays[attempt]}초 후 재시도...")
        
        print(f"    [최종 실패] 3번 시도 후에도 실패: {str(last_error)}")
        return []

    def _channels_with_retry(self, channel_ids):
        """네트워크 오류 방지를 위한 재시도 기능이 탑재된 안정적인 채널 API 호출 함수"""
        import time
        
        if self.is_cancelled: return []
        
        retry_delays = [2, 4]  # 2초, 4초 간격으로 재시도
        last_error = None
        
        for attempt in range(3):  # 최대 3번 시도
            try:
                response = self.youtube.channels().list(
                    part='statistics',
                    id=','.join(channel_ids)
                ).execute()
                return response.get('items', [])
            except Exception as e:
                last_error = e
                print(f"    [재시도 {attempt + 1}/3] 채널 API 호출 실패: {str(e)}")
                
                if attempt < 2:  # 마지막 시도가 아니면 대기
                    time.sleep(retry_delays[attempt])
                    print(f"    {retry_delays[attempt]}초 후 재시도...")
        
        print(f"    [최종 실패] 3번 시도 후에도 실패: {str(last_error)}")
        return []

    def _get_published_after_date(self, period, search_time=None):
        """업로드 기간에 따른 날짜 계산 (SimilarVideoWorker용)"""
        # 내부 코드값 기반 처리 (UI/로직 분리)
        if period == 'all':
            return None
        
        # 검색 시점을 고정하여 일관성 개선
        now = search_time or datetime.now(timezone.utc)
        # 내부 코드값을 일수로 매핑
        days_map = {
            'day': 1, 
            'week': 7, 
            'month': 30, 
            'month2': 60, 
            'month3': 90, 
            'month6': 180, 
            'year': 365
        }
        days = days_map.get(period, 60)  # 기본값: 2개월
        
        if days == 0:
            return None
        
        # API가 요구하는 'RFC 3339' 형식의 문자열로 변환하여 반환
        published_after = (now - timedelta(days=days)).isoformat().replace('+00:00', 'Z')
        return published_after

    def _extract_core_keywords(self, title):
        """제목에서 핵심 키워드만 추출하여 노이즈 제거"""
        import re
        
        if not title:
            return title
            
        # 노이즈 패턴들 (감정적 표현, 일반적 단어 등)
        noise_patterns = [
            r'[😀-🙏]',  # 이모지
            r'#\w+',     # 해시태그
            r'\b(때|할때|하는|이런|저런|그런|어떤|무슨|진짜|정말|너무|완전|엄청|대박|최고|최악)\b',
            r'\b(외로울|슬픈|기쁜|화난|답답한|심심한|재미있는|지루한)\b',
            r'\b(시작|끝|마지막|처음|중간|다음|이전)\b',
            r'\b(한|두|세|네|다섯|여섯|일곱|여덟|아홉|열)\b',
            r'\b(오늘|어제|내일|지금|나중에|요즘|최근)\b'
        ]
        
        filtered_title = title
        for pattern in noise_patterns:
            filtered_title = re.sub(pattern, '', filtered_title)
        
        # 연속된 공백 정리
        filtered_title = re.sub(r'\s+', ' ', filtered_title).strip()
        
        # 너무 짧아지면 원본 사용
        if len(filtered_title) < len(title) * 0.3:
            return title
            
        return filtered_title if filtered_title else title

    def _detect_content_language(self, title, channel_title):
        """콘텐츠의 주요 언어 감지"""
        import re
        
        # 언어별 패턴
        korean_pattern = r'[가-힣]'
        chinese_pattern = r'[\u4e00-\u9fff\u3400-\u4dbf]'
        japanese_pattern = r'[\u3040-\u309f\u30a0-\u30ff]'
        
        combined_text = f"{title} {channel_title}"
        
        # 각 언어별 문자 수 계산
        korean_chars = len(re.findall(korean_pattern, combined_text))
        chinese_chars = len(re.findall(chinese_pattern, combined_text))
        japanese_chars = len(re.findall(japanese_pattern, combined_text))
        
        total_asian_chars = korean_chars + chinese_chars + japanese_chars
        
        if total_asian_chars == 0:
            return 'english'  # 아시아 문자가 없으면 영어로 분류
        
        # 가장 많은 비율을 차지하는 언어 반환
        if korean_chars >= chinese_chars and korean_chars >= japanese_chars:
            return 'korean'
        elif chinese_chars >= japanese_chars:
            return 'chinese'
        else:
            return 'japanese'

    def _is_similar_language_content(self, title, channel_title, reference_language):
        """
        (할당량 부족 시 관대한 로직) 원본 영상과 유사한 언어의 콘텐츠인지 판단합니다.
        """
        # 현재 콘텐츠의 언어를 감지합니다.
        content_language = self._detect_content_language(title, channel_title)
        
        # 원본 영상의 언어와 현재 콘텐츠의 언어가 정확히 일치하는 경우
        if content_language == reference_language:
            return True
        
        # API 할당량 부족으로 검색 결과가 적을 때 더 관대하게 처리
        # 한국어 기준으로 영어도 허용 (비즈니스/마케팅 콘텐츠는 영어 용어 많이 사용)
        if reference_language == 'ko' and content_language in ['en', 'korean']:
            return True
        
        # 영어 기준으로 한국어도 허용
        if reference_language == 'en' and content_language in ['ko', 'korean']:
            return True
        
        # 그 외 모든 경우는 관련 없는 언어로 판단
        return False

    def _format_duration(self, duration):
        """PT 형식의 duration을 읽기 쉬운 형식으로 변환"""
        import re
        
        if not duration:
            return "0:00"
            
        # PT1H2M3S 형식 파싱
        pattern = r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?'
        match = re.match(pattern, duration)
        
        if not match:
            return "0:00"
            
        hours = int(match.group(1)) if match.group(1) else 0
        minutes = int(match.group(2)) if match.group(2) else 0
        seconds = int(match.group(3)) if match.group(3) else 0
        
        if hours > 0:
            return f"{hours}:{minutes:02d}:{seconds:02d}"
        else:
            return f"{minutes}:{seconds:02d}"

    def _sort_results(self, results):
        """결과를 정렬 방식에 따라 정렬 (Worker와 동일한 로직)"""
        if self.sort_by == 'viewCount':
            # 조회수 기준 내림차순 정렬
            results.sort(key=lambda x: x['view_count'], reverse=True)
        elif self.sort_by == 'engagement_rate':
            # 참여도 기준 내림차순 정렬 (좋아요/조회수)
            results.sort(key=lambda x: x['engagement_rate'], reverse=True)
        elif self.sort_by == 'reaction_rate':
            # 반응도 기준 내림차순 정렬 (댓글/조회수)
            results.sort(key=lambda x: x['reaction_rate'], reverse=True)
        elif self.sort_by == 'date_desc':
            # 최신순 (날짜 내림차순)
            results.sort(key=lambda x: x['published_at_full'], reverse=True)
        elif self.sort_by == 'date_asc':
            # 오래된순 (날짜 오름차순)
            results.sort(key=lambda x: x['published_at_full'], reverse=False)
        
        return results

    def run(self):
        """
        [NEW AND IMPROVED VERSION]
        - Task 1 Fix: Replaced parallel processing with a stable sequential process to prevent crashes.
        - Task 2 Fix: Implemented a weighted progress system for smooth UI updates.
        - Task 3 Fix: Checks the `self.is_cancelled` flag in loops to allow graceful termination.
        - Task 4 Fix: Added progress updates for final filtering loop (80% -> 100%).
        - Task 5 Fix: Prioritized channel search and increased maxResults to 40.
        """
        try:
            # === Step 0: Setup for Progress Calculation ===
            progress = 0
            weights = {'initial': 10, 'collect': 30, 'details': 40, 'finalize': 20}

            # === Step 1: Core Data Extraction & Preparation (Progress: 0% -> 10%) ===
            if self.is_cancelled: return
            self.progress.emit(self._("원본 영상 분석 중... ({}%)").format(progress))
            self.youtube = build('youtube', 'v3', developerKey=self.api_key)

            video_id = self._extract_video_id_from_url(self.video_url)
            if not video_id: raise ValueError(self._("올바른 유튜브 영상 URL을 입력해주세요."))

            original_video_data = self._videos_with_retry([video_id])
            if not original_video_data: raise ValueError(self._("해당 영상을 찾을 수 없습니다."))

            original_video = original_video_data[0]
            original_snippet = original_video['snippet']
            original_title = original_snippet.get('title', '')
            original_description = original_snippet.get('description', '')
            original_tags = original_snippet.get('tags', [])
            original_channel_id = original_snippet.get('channelId', '')
            
            # 해시태그 추출 (제목 + 설명에서)
            import re
            original_hashtags = []
            original_hashtags.extend(re.findall(r"#(\w+)", original_title))
            original_hashtags.extend(re.findall(r"#(\w+)", original_description))
            original_hashtags = list(set(original_hashtags))  # 중복 제거
            
            # topicDetails 정보 추출
            topic_details = original_video.get('topicDetails', {})
            topic_ids = topic_details.get('relevantTopicIds', [])

            search_time = datetime.now(timezone.utc)
            published_after = self._get_published_after_date(self.upload_period, search_time)
            
            # 언어 감지
            try:
                actual_language = self._detect_content_language(original_title, original_snippet.get('channelTitle', ''))
            except Exception as e:
                actual_language = 'ko'
            
            default_language = original_snippet.get('defaultAudioLanguage', 'ko')
            
            # 안전한 언어 코드 매핑
            language_to_region = {'ko': 'KR', 'en': 'US', 'ja': 'JP', 'zh': 'CN', 'es': 'ES', 'fr': 'FR', 'de': 'DE', 'ru': 'RU', 'pt': 'BR', 'it': 'IT'}
            safe_language_codes = {'ko', 'en', 'ja', 'zh', 'es', 'fr', 'de', 'ru', 'pt', 'it'}
            
            detected_lang = actual_language if actual_language in safe_language_codes else 'ko'
            
            api_lang = default_language[:2] if default_language and len(default_language) >= 2 else 'ko'
            if api_lang in safe_language_codes and actual_language == 'ko' and api_lang != 'ko':
                detected_lang = 'ko'
            
            if detected_lang not in safe_language_codes:
                detected_lang = 'ko'
            
            self.region_code = language_to_region.get(detected_lang, 'KR')
            self.language_code = detected_lang

            # 동영상 길이 감지 및 필터 조정
            original_duration = original_video.get('contentDetails', {}).get('duration', 'PT0S')
            duration_match = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', original_duration)
            original_seconds = 0
            if duration_match:
                h, m, s = [int(g or 0) for g in duration_match.groups()]
                original_seconds = h * 3600 + m * 60 + s
            
            is_original_short = original_seconds < 60
            
            adjusted_video_duration = self.video_duration
            
            if self.video_duration == 'any':
                adjusted_video_duration = 'any'
            elif is_original_short and self.video_duration == 'long':
                adjusted_video_duration = 'any'
            elif not is_original_short and self.video_duration == 'short':
                adjusted_video_duration = 'any'

            progress = weights['initial']
            if self.is_cancelled: return
            self.progress.emit(self._("후보 영상 수집 준비... ({}%)").format(progress))

            # === Step 2: [Task 1 Fix] Safe Sequential Candidate Collection (Progress: 10% -> 40%) ===
            candidate_video_ids = set()
            # [Task 5 Fix] Prioritize Channel Search and Increase maxResults to 40
            search_tasks = [
                {'channelId': original_channel_id, 'maxResults': 40},  # Prioritized and increased from 20 to 40
                {'q': original_title, 'maxResults': 25},
                {'q': ' '.join(original_tags[:3]) if original_tags else original_snippet.get('channelTitle', ''), 'maxResults': 20},
                {'q': ' '.join([f"#{tag}" for tag in original_hashtags[:3]]) if original_hashtags else ' '.join([word for word in original_title.split() if len(word) > 2][:3]), 'maxResults': 15},
                {'topicId': topic_ids[0] if topic_ids else None, 'videoCategoryId': original_snippet.get('categoryId', '22') if not topic_ids else None, 'maxResults': 15}
            ]

            # Common parameters
            common_params = {
                'part': 'snippet', 
                'type': 'video', 
                'relevanceLanguage': self.language_code, 
                'regionCode': self.region_code
            }
            
            if adjusted_video_duration != 'any':
                common_params['videoDuration'] = adjusted_video_duration
            if published_after:
                common_params['publishedAfter'] = published_after

            total_search_tasks = len(search_tasks)
            for i, task_params in enumerate(search_tasks):
                if self.is_cancelled: return

                # [Task 2 Fix] Calculate and emit progress for each search task
                progress_step = weights['collect'] / total_search_tasks
                current_progress = int(progress + (i * progress_step))
                self.progress.emit(self._("후보 영상 수집 중({}/{})... ({}%)").format(i + 1, total_search_tasks, current_progress))

                # Merge common and specific parameters
                params = {**common_params, **task_params}
                # Remove None values
                params = {k: v for k, v in params.items() if v is not None}

                search_results = self._search_with_retry(params)
                for item in search_results:
                    vid_id = item.get('id', {}).get('videoId')
                    if vid_id: candidate_video_ids.add(vid_id)

            if self.is_cancelled: return
            progress += weights['collect']

            # === Step 3: Detailed Info Collection (Progress: 40% -> 80%) ===
            if not candidate_video_ids:
                self.finished.emit({'results': [], 'total_count': 0})
                return

            candidate_list = list(candidate_video_ids)
            all_video_details = []
            batch_size = 50
            total_batches = (len(candidate_list) + batch_size - 1) // batch_size

            for i in range(total_batches):
                if self.is_cancelled: return

                # [Task 2 Fix] Calculate and emit progress for each batch
                progress_step = weights['details'] / total_batches
                current_progress = int(progress + (i * progress_step))
                self.progress.emit(self._("영상 상세정보 수집 중({}/{})... ({}%)").format(i + 1, total_batches, current_progress))

                batch = candidate_list[i*batch_size : (i+1)*batch_size]
                batch_results = self._videos_with_retry(batch)
                all_video_details.extend(batch_results)

            if self.is_cancelled: return
            progress += weights['details']

            # === Step 4: Scoring & Final Filtering (Progress: 80% -> 100%) ===
            # 스코어링
            scored_videos = []
            title_keywords = set(word for word in original_title.split() if len(word) > 2)

            for video in all_video_details:
                score = 0
                snippet = video.get('snippet', {})
                
                # 채널 일치 보너스
                if snippet.get('channelId') == original_channel_id: 
                    score += 30
                
                # 태그 일치 보너스
                video_tags = snippet.get('tags', [])
                if video_tags and original_tags:
                    common_tags = set(original_tags) & set(video_tags)
                    score += len(common_tags) * 15

                # 해시태그 일치 보너스
                video_description = snippet.get('description', '')
                video_hashtags = re.findall(r"#(\w+)", video_description)
                if video_hashtags and original_hashtags:
                    common_hashtags = set(original_hashtags) & set(video_hashtags)
                    score += len(common_hashtags) * 8

                # 제목 키워드 일치 보너스
                video_title = snippet.get('title', '').lower()
                for keyword in title_keywords:
                    if keyword.lower() in video_title:
                        score += 5
                
                video['similarity_score'] = score
                scored_videos.append(video)
            
            # 점수 높은 순으로 정렬
            scored_videos.sort(key=lambda x: x['similarity_score'], reverse=True)

            # 채널 구독자 수 정보 가져오기 (상위 50개만)
            final_channel_ids = list(set(v['snippet']['channelId'] for v in scored_videos[:50]))
            subscriber_counts = {}
            
            for i in range(0, len(final_channel_ids), 50):
                batch = final_channel_ids[i:i+50]
                channel_results = self._channels_with_retry(batch)
                for item in channel_results:
                    channel_id = item['id']
                    subscriber_count = int(item['statistics'].get('subscriberCount', 0))
                    subscriber_counts[channel_id] = subscriber_count
            
            # 원본 언어 감지
            original_lang = self._detect_content_language(original_title, original_snippet.get('channelTitle', ''))

            # 최종 필터링 - [Task 4 Fix] Added progress updates for finalize stage (80% -> 100%)
            final_results_raw = []
            total_videos_to_process = min(50, len(scored_videos))
            
            for idx, video in enumerate(scored_videos[:50]):  # 상위 50개만 처리
                if self.is_cancelled: return

                # [Task 4 Fix] Calculate and emit progress for final filtering loop
                finalize_progress = (idx + 1) / total_videos_to_process * weights['finalize']
                current_progress = int(progress + finalize_progress)
                self.progress.emit(self._("최종 필터링 적용 중({}/{})... ({}%)").format(idx + 1, total_videos_to_process, current_progress))

                sub_count = subscriber_counts.get(video['snippet']['channelId'], 0)
                view_count = int(video.get('statistics', {}).get('viewCount', 0))
                
                # 언어 필터
                is_lang_match = self._is_similar_language_content(
                    video['snippet']['title'], 
                    video['snippet']['channelTitle'], 
                    original_lang
                )
                
                if not is_lang_match:
                    continue

                # 동영상 길이 필터
                if adjusted_video_duration == 'any':
                    duration_str = video.get('contentDetails', {}).get('duration', 'PT0S')
                    total_seconds = 0
                    match = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', duration_str)
                    if match:
                        h, m, s = [int(g or 0) for g in match.groups()]
                        total_seconds = h * 3600 + m * 60 + s

                    duration_match = False
                    if self.video_duration == 'any':
                        duration_match = True
                    elif self.video_duration == 'short' and total_seconds < 60:
                        duration_match = True
                    elif self.video_duration == 'long' and total_seconds >= 60:
                        duration_match = True

                    if not duration_match:
                        continue

                # 구독자/조회수 필터
                if not (view_count >= self.min_views and 
                        (self.max_subs == 0 or sub_count <= self.max_subs)):
                    continue
                    
                # 최종 결과에 추가
                final_results_raw.append({
                    'video_id': video['id'], 
                    'published_at': video['snippet']['publishedAt'].split('T')[0],
                    'published_at_full': video['snippet']['publishedAt'], 
                    'view_count': view_count,
                    'like_count': int(video.get('statistics', {}).get('likeCount', 0)), 
                    'comment_count': int(video.get('statistics', {}).get('commentCount', 0)),
                    'duration': video.get('contentDetails', {}).get('duration', 'PT0M0S'), 
                    'duration_formatted': self._format_duration(video.get('contentDetails', {}).get('duration', 'PT0M0S')),
                    'title': video['snippet']['title'], 
                    'channel_title': video['snippet']['channelTitle'],
                    'channel_id': video['snippet']['channelId'], 
                    'subscriber_count': sub_count,
                    'engagement_rate': int(video.get('statistics', {}).get('likeCount', 0)) / max(view_count, 1),
                    'reaction_rate': int(video.get('statistics', {}).get('commentCount', 0)) / max(view_count, 1),
                    'license': video.get('status', {}).get('license', 'youtube'),
                    'thumbnail_url': video['snippet']['thumbnails'].get('medium', {}).get('url', ''),
                    'similarity_score': video['similarity_score']
                })

            if self.is_cancelled: return
            final_results_sorted = self._sort_results(final_results_raw)

            self.progress.emit(self._("검색 완료! 최종 {}개 영상 발견 (100%)").format(len(final_results_sorted)))
            self.finished.emit({'results': final_results_sorted, 'total_count': len(final_results_sorted)})

        except Exception as e:
            import traceback
            import json
            from googleapiclient.errors import HttpError
            
            if isinstance(e, HttpError):
                try:
                    error_details = json.loads(e.content.decode('utf-8'))['error']
                    reason = error_details.get('errors', [{}])[0].get('reason', 'unknown')
                    if reason == 'quotaExceeded': self.error.emit("QUOTA_EXCEEDED")
                    elif reason == 'keyInvalid': self.error.emit("KEY_INVALID")
                    else: self.error.emit(self._("유튜브 API 오류: ") + error_details.get('message', str(e)))
                except:
                    error_message = self._("유튜브 API 오류: ") + str(e)
                    self.error.emit(error_message)
            else:
                # 일반 오류의 경우
                error_message = self._("오류 발생: ") + str(e)
                self.error.emit(error_message)
            
            traceback.print_exc()
    
    def _get_video_details_parallel(self, video_ids):
        """병렬로 비디오 상세 정보 가져오기"""
        from concurrent.futures import ThreadPoolExecutor, as_completed
        
        def get_video_batch(batch_ids):
            try:
                youtube = build('youtube', 'v3', developerKey=self.api_key)
                response = youtube.videos().list(
                    part='snippet,statistics,contentDetails,status',
                    id=','.join(batch_ids)
                ).execute()
                return response.get('items', [])
            except Exception as e:
                return []
        
        all_video_details = []
        batch_size = 50
        
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = []
            for i in range(0, len(video_ids), batch_size):
                batch = video_ids[i:i+batch_size]
                future = executor.submit(get_video_batch, batch)
                futures.append(future)
            
            for future in as_completed(futures):
                batch_results = future.result()
                all_video_details.extend(batch_results)
        
        return all_video_details

    def _get_channel_stats_parallel(self, channel_ids):
        """병렬로 채널 통계 가져오기"""
        from concurrent.futures import ThreadPoolExecutor, as_completed
        
        def get_channel_batch(batch_ids):
            try:
                youtube = build('youtube', 'v3', developerKey=self.api_key)
                response = youtube.channels().list(
                    part='statistics',
                    id=','.join(batch_ids)
                ).execute()
                
                subscriber_counts = {}
                for item in response.get('items', []):
                    channel_id = item['id']
                    subscriber_count = int(item['statistics'].get('subscriberCount', 0))
                    subscriber_counts[channel_id] = subscriber_count
                
                return subscriber_counts
            except Exception as e:
                return {}
        
        all_subscriber_counts = {}
        batch_size = 50
        
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = []
            for i in range(0, len(channel_ids), batch_size):
                batch = channel_ids[i:i+batch_size]
                future = executor.submit(get_channel_batch, batch)
                futures.append(future)
            
            for future in as_completed(futures):
                batch_results = future.result()
                all_subscriber_counts.update(batch_results)
        
        return all_subscriber_counts


class Worker(QThread):
    """메인 검색 워커 클래스"""
    progress = Signal(str)
    finished = Signal(dict)
    error = Signal(str)

    def __init__(self, search_keyword, upload_period, min_views, sort_by, video_duration, max_subs, api_key, translator_func, similar_video_url=None, analyze_contribution=False):
        super().__init__()
        self.api_key = api_key
        self.search_keyword = search_keyword
        self.upload_period = upload_period
        self.min_views = min_views
        self.sort_by = sort_by
        self.video_duration = video_duration
        self.max_subs = max_subs
        self.similar_video_url = similar_video_url
        self.analyze_contribution = analyze_contribution
        self.youtube = None
        self._channel_cache = {}
        self._video_cache = {}
        self._ = translator_func

    def _(self, text):
        """번역 함수"""
        return text

    def run(self):
        try:
            # 키워드 검색 전용 (유사 영상 검색은 SimilarVideoWorker에서 처리)
            if not self.search_keyword:
                raise ValueError(self._("검색 키워드를 입력해주세요."))
            if not self.api_key:
                raise ValueError(self._("API key가 올바르지 않습니다."))
            
            # --- 이 아래는 기존의 '키워드 검색' 로직입니다 ---
            self.youtube = build('youtube', 'v3', developerKey=self.api_key)
            
            search_time = datetime.now()
            published_after = self._get_published_after_date(self.upload_period, search_time)
            
            self.progress.emit(self._("동영상 검색 시작..."))
            
            max_pages = 6
            api_order = self._get_api_order(self.sort_by)
            
            all_video_ids = self._search_videos_parallel(api_order, published_after, max_pages)
            all_video_ids = list(dict.fromkeys(all_video_ids))

            if not all_video_ids:
                self.finished.emit({'results': [], 'total_count': 0})
                return

            self.progress.emit(self._("총 {}개 동영상 정보 수집 중...").format(len(all_video_ids)))
            all_video_details = self._get_video_details_parallel(all_video_ids)
            
            channel_ids = list(set(v['snippet']['channelId'] for v in all_video_details))
            subscriber_counts = self._get_channel_stats_parallel(channel_ids)

            # 기여도 분석은 주문형으로 변경 - API 절약을 위해 제거

            self.progress.emit(self._("필터링 및 정렬 중..."))
            
            # 언어 기준 설정 (UI에서 설정된 언어 사용)
            reference_language = 'ko'  # 기본값, 필요시 UI에서 가져오도록 수정 가능
            
            results = []
            for video in all_video_details:
                view_count = int(video['statistics'].get('viewCount', 0))
                like_count = int(video['statistics'].get('likeCount', 0))
                comment_count = int(video['statistics'].get('commentCount', 0))
                duration = video['contentDetails'].get('duration', 'PT0M0S')
                channel_id = video['snippet']['channelId']
                subscriber_count = subscriber_counts.get(channel_id, 0)
                license_info = video.get('status', {}).get('license', 'youtube')
                engagement_rate = like_count / max(view_count, 1)
                reaction_rate = comment_count / max(view_count, 1)

                # 엄격한 언어 필터 적용
                is_lang_match = self._is_similar_language_content(
                    video['snippet']['title'], 
                    video['snippet']['channelTitle'], 
                    reference_language
                )

                if (view_count >= self.min_views and 
                    (self.max_subs == 0 or subscriber_count <= self.max_subs) and
                    is_lang_match):
                    
                    result_item = {
                        'video_id': video['id'],
                        'published_at': video['snippet']['publishedAt'].split('T')[0],
                        'published_at_full': video['snippet']['publishedAt'],
                        'view_count': view_count,
                        'like_count': like_count,
                        'comment_count': comment_count,
                        'duration': duration,
                        'duration_formatted': self._format_duration(duration),
                        'title': video['snippet']['title'],
                        'channel_title': video['snippet']['channelTitle'],
                        'channel_id': channel_id,
                        'subscriber_count': subscriber_count,
                        'engagement_rate': engagement_rate,
                        'reaction_rate': reaction_rate,
                        'license': license_info,
                        'thumbnail_url': video['snippet']['thumbnails'].get('medium', {}).get('url', 
                                       video['snippet']['thumbnails'].get('default', {}).get('url', ''))
                    }
            
                    # 기여도 분석은 주문형으로 변경 - 기본 검색에서는 제거
                    
                    results.append(result_item)
            
            results = self._sort_results(results)
            
            self.progress.emit(self._("완료!"))
            self.finished.emit({'results': results, 'total_count': len(results)})

        except Exception as e:
            import traceback
            import json
            from googleapiclient.errors import HttpError
            
            # HttpError 구체적 파싱 및 오류 코드 emit
            if isinstance(e, HttpError):
                try:
                    error_details = json.loads(e.content.decode('utf-8'))['error']
                    reason = error_details.get('errors', [{}])[0].get('reason', 'unknown')
                    
                    if reason == 'quotaExceeded':
                        self.error.emit("QUOTA_EXCEEDED")  # 특정 오류 코드 emit
                    elif reason == 'keyInvalid':
                        self.error.emit("KEY_INVALID")     # 특정 오류 코드 emit
                    else:
                        error_message = self._("유튜브 API 오류: ") + error_details.get('message', str(e))
                        self.error.emit(error_message)
                except:
                    error_message = self._("유튜브 API 오류: ") + str(e)
                    self.error.emit(error_message)
            else:
                # 일반 오류의 경우
                error_message = self._("오류 발생: ") + str(e)
                self.error.emit(error_message)
            
            traceback.print_exc()
    
    def _search_videos_parallel(self, api_order, published_after, max_pages):
        """병렬로 비디오 검색 (할당량 소진 시 부분 결과 반환)"""
        quota_exhausted = False  # 할당량 소진 플래그
        
        def search_page(page_info, translator):
            nonlocal quota_exhausted
            page_token, page_num = page_info
            try:
                # 각 스레드에서 별도의 YouTube 클라이언트 사용
                youtube = build('youtube', 'v3', developerKey=self.api_key)
                
                # 언어-지역 매핑 (키워드 검색 정확도 향상)
                language_to_region = {
                    'ko': 'KR', 'en': 'US', 'ja': 'JP', 'zh': 'CN', 'es': 'ES',
                    'fr': 'FR', 'de': 'DE', 'ru': 'RU', 'pt': 'BR', 'it': 'IT'
                }
                
                search_params = {
                    'q': self.search_keyword,
                    'part': 'snippet',
                    'type': 'video',
                    'order': api_order,
                    'videoDuration': self.video_duration,
                    'maxResults': 50,
                    'relevanceLanguage': 'ko',  # 기본 한국어
                    'regionCode': language_to_region.get('ko', 'KR')  # 기본 한국
                }
                
                if published_after:
                    search_params['publishedAfter'] = published_after
                    
                if page_token:
                    search_params['pageToken'] = page_token
                
                response = youtube.search().list(**search_params).execute()
                
                videos = [item['id']['videoId'] for item in response.get('items', [])]
                next_token = response.get('nextPageToken')
                
                self.progress.emit(self._("동영상 검색 중... ({}/{})").format(page_num + 1, max_pages))
                
                return videos, next_token
                
            except Exception as e:
                from googleapiclient.errors import HttpError
                import json
                
                # HttpError인 경우 할당량 소진 여부 확인
                if isinstance(e, HttpError):
                    try:
                        error_details = json.loads(e.content.decode('utf-8'))['error']
                        reason = error_details.get('errors', [{}])[0].get('reason', 'unknown')
                        
                        if reason == 'quotaExceeded':
                            quota_exhausted = True
                            self.progress.emit(self._("API 할당량 소진 - 수집된 결과만 표시합니다"))
                            return [], None  # 빈 결과 반환하여 더 이상 검색하지 않음
                    except:
                        pass
                
                return [], None
        
        all_video_ids = []
        page_tokens = [None]  # 첫 번째 페이지는 토큰이 없음
        
        # 첫 번째 페이지를 먼저 가져와서 페이지 토큰들을 수집
        first_videos, next_token = search_page((None, 0), self._)
        all_video_ids.extend(first_videos)
        
        # 할당량이 소진되었으면 첫 페이지 결과만 반환
        if quota_exhausted:
            if all_video_ids:
                self.progress.emit(self._("할당량 소진으로 {}개 결과만 표시됩니다").format(len(all_video_ids)))
            return all_video_ids
        
        # 나머지 페이지 토큰들을 순차적으로 수집
        current_token = next_token
        for page_num in range(1, max_pages):
            if not current_token or quota_exhausted:
                break
            page_tokens.append(current_token)
            
            # 다음 토큰을 얻기 위해 페이지를 하나씩 가져옴 (YouTube API 특성상 순차적 필요)
            _, current_token = search_page((current_token, page_num), self._)
            
            # 할당량 소진 시 조기 종료
            if quota_exhausted:
                break
        
        # 수집된 토큰들로 병렬 처리 (첫 번째 제외)
        if len(page_tokens) > 1 and not quota_exhausted:
            with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
                page_infos = [(token, idx) for idx, token in enumerate(page_tokens[1:], 1)]
                future_to_page = {executor.submit(search_page, page_info, self._): page_info for page_info in page_infos}
                
                for future in concurrent.futures.as_completed(future_to_page):
                    try:
                        videos, _ = future.result(timeout=30)  # 30초 타임아웃
                        all_video_ids.extend(videos)
                        
                        # 할당량 소진 시 조기 종료
                        if quota_exhausted:
                            break
                    except Exception as e:
                        continue
        
        # 최종 결과 반환 (할당량 소진 여부와 관계없이)
        if quota_exhausted and all_video_ids:
            self.progress.emit(self._("할당량 소진으로 {}개 결과만 표시됩니다").format(len(all_video_ids)))
        
        return all_video_ids
    
    def _get_video_details_parallel(self, video_ids):
        """병렬로 비디오 상세 정보 수집 (할당량 소진 시 부분 결과 반환)"""
        quota_exhausted = False  # 할당량 소진 플래그
        
        def get_video_batch(batch_ids, translator):
            nonlocal quota_exhausted
            try:
                # 각 스레드에서 별도의 YouTube 클라이언트 사용
                youtube = build('youtube', 'v3', developerKey=self.api_key)
                return youtube.videos().list(
                    part='snippet,statistics,contentDetails,status',
                    id=','.join(batch_ids)
                ).execute().get('items', [])
                
            except Exception as e:
                from googleapiclient.errors import HttpError
                import json
                
                # HttpError인 경우 할당량 소진 여부 확인
                if isinstance(e, HttpError):
                    try:
                        error_details = json.loads(e.content.decode('utf-8'))['error']
                        reason = error_details.get('errors', [{}])[0].get('reason', 'unknown')
                        
                        if reason == 'quotaExceeded':
                            quota_exhausted = True
                            self.progress.emit(self._("API 할당량 소진 - 수집된 정보만 표시합니다"))
                    except:
                        pass
                
                return []
        
        all_video_details = []
        batch_size = 50
        batches = [video_ids[i:i + batch_size] for i in range(0, len(video_ids), batch_size)]
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
            future_to_batch = {executor.submit(get_video_batch, batch, self._): idx for idx, batch in enumerate(batches)}
            
            for future in concurrent.futures.as_completed(future_to_batch):
                batch_idx = future_to_batch[future]
                try:
                    video_details = future.result(timeout=30)  # 30초 타임아웃
                    all_video_details.extend(video_details)
                    
                    # 진행률 업데이트
                    completed_batches = batch_idx + 1
                    total_batches = len(batches)
                    self.progress.emit(self._("동영상 정보 수집 중... ({}/{})").format(completed_batches, total_batches))
                    
                    # 할당량 소진 시 조기 종료
                    if quota_exhausted:
                        self.progress.emit(self._("할당량 소진으로 {}개 영상 정보만 수집됨").format(len(all_video_details)))
                        break
                        
                except Exception as e:
                    continue
        
        return all_video_details
    
    def _get_channel_stats_parallel(self, channel_ids):
        """병렬로 채널 통계 정보 수집 (캐싱 적용)"""
        # 캐시에서 이미 있는 채널 정보 확인
        cached_channels = {}
        uncached_channel_ids = []
        
        for channel_id in channel_ids:
            if channel_id in self._channel_cache:
                cached_channels[channel_id] = self._channel_cache[channel_id]
            else:
                uncached_channel_ids.append(channel_id)
        
        # 캐시되지 않은 채널들만 API 호출
        if uncached_channel_ids:
            def get_channel_batch(batch_ids, translator):
                try:
                    # 각 스레드에서 별도의 YouTube 클라이언트 사용
                    youtube = build('youtube', 'v3', developerKey=self.api_key)
                    return youtube.channels().list(
                        part='statistics',
                        id=','.join(batch_ids)
                    ).execute().get('items', [])
                except Exception as e:
                    return []
            
            batch_size = 50
            batches = [uncached_channel_ids[i:i + batch_size] for i in range(0, len(uncached_channel_ids), batch_size)]
            
            all_channel_stats = []
            with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
                future_to_batch = {executor.submit(get_channel_batch, batch, self._): batch for batch in batches}
                
                for future in concurrent.futures.as_completed(future_to_batch):
                    try:
                        channel_stats = future.result(timeout=30)  # 30초 타임아웃
                        all_channel_stats.extend(channel_stats)
                    except Exception as e:
                        continue
            
            # 새로 가져온 채널 정보를 캐시에 저장
            for item in all_channel_stats:
                subscriber_count = int(item['statistics'].get('subscriberCount', 0))
                self._channel_cache[item['id']] = subscriber_count
                cached_channels[item['id']] = subscriber_count
        
        return cached_channels

    def _get_api_order(self, sort_by):
        """YouTube API의 order 파라미터를 반환 - 일관성을 위해 relevance 사용"""
        # 문제 해결: 모든 정렬에 대해 relevance 사용하여 일관된 결과 확보
        # YouTube API의 정렬은 페이지별로 작동하므로 여러 페이지 합칠 때 순서가 뒤바뀜
        # Python에서만 정렬하여 정확한 순서 보장
        return 'relevance'

    def _sort_results(self, results):
        """결과를 정렬 방식에 따라 정렬"""
        if self.sort_by == 'viewCount':
            # 조회수 기준 내림차순 정렬
            results.sort(key=lambda x: x['view_count'], reverse=True)
        elif self.sort_by == 'engagement_rate':
            # 참여도 기준 내림차순 정렬 (좋아요/조회수)
            results.sort(key=lambda x: x['engagement_rate'], reverse=True)
        elif self.sort_by == 'reaction_rate':
            # 반응도 기준 내림차순 정렬 (댓글/조회수)
            results.sort(key=lambda x: x['reaction_rate'], reverse=True)
        elif self.sort_by == 'date_desc':
            # 최신순 (날짜 내림차순)
            results.sort(key=lambda x: x['published_at_full'], reverse=True)
        elif self.sort_by == 'date_asc':
            # 오래된순 (날짜 오름차순)
            results.sort(key=lambda x: x['published_at_full'], reverse=False)
        
        return results

    def _format_duration(self, duration):
        """ISO 8601 duration을 사람이 읽기 쉬운 형식으로 변환 (롱폼/쇼츠 구분)"""
        # PT15M33S -> 15:33 (쇼츠)
        # PT1H23M45S -> 83:45 (롱폼 - 총 분:초로 표시)
        # PT45S -> 0:45 (쇼츠)
        
        match = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', duration)
        if not match:
            return "0:00"
        
        hours = int(match.group(1) or 0)
        minutes = int(match.group(2) or 0)
        seconds = int(match.group(3) or 0)
        
        # 1시간 이상이면 총 분:초로 표시 (예: 1시간 11분 32초 → 71:32)
        if hours > 0:
            total_minutes = hours * 60 + minutes
            return f"{total_minutes}:{seconds:02d}"
        else:
            # 1시간 미만이면 분:초로 표시
            if minutes == 0 and seconds > 0:
                return f"0:{seconds:02d}"
            else:
                return f"{minutes}:{seconds:02d}"

    def _get_published_after_date(self, period, search_time=None):
        # 내부 코드값 기반 처리 (UI/로직 분리)
        if period == 'all':
            return None
        
        # 검색 시점을 고정하여 일관성 개선
        now = search_time or datetime.now()
        # 내부 코드값을 일수로 매핑
        days_map = {
            'day': 1, 
            'week': 7, 
            'month': 30, 
            'month2': 60, 
            'month3': 90, 
            'month6': 180, 
            'year': 365
        }
        days = days_map.get(period, 60)  # 기본값: 2개월
        
        if days == 0:
            return None
        
        published_after = (now - timedelta(days=days)).isoformat("T") + "Z"
        return published_after

    def _detect_content_language(self, title, channel_title):
        """제목과 채널명을 기반으로 콘텐츠 언어를 감지합니다."""
        text = f"{title} {channel_title}".lower()
        
        # 한국어 패턴 감지
        korean_patterns = [
            r'[가-힣]',  # 한글
            r'[\u3131-\u3163]',  # 한글 자모
        ]
        
        # 영어 패턴 감지 (알파벳만으로 구성된 단어들)
        english_patterns = [
            r'\b[a-z]+\b',  # 영어 단어
        ]
        
        # 일본어 패턴 감지
        japanese_patterns = [
            r'[\u3040-\u309F]',  # 히라가나
            r'[\u30A0-\u30FF]',  # 가타카나
        ]
        
        # 중국어 패턴 감지
        chinese_patterns = [
            r'[\u4E00-\u9FFF]',  # 한자 (중국어/일본어 공통이지만 중국어로 분류)
        ]
        
        import re
        
        # 각 언어 패턴 매칭 개수 계산
        korean_matches = sum(len(re.findall(pattern, text)) for pattern in korean_patterns)
        english_matches = sum(len(re.findall(pattern, text)) for pattern in english_patterns)
        japanese_matches = sum(len(re.findall(pattern, text)) for pattern in japanese_patterns)
        chinese_matches = sum(len(re.findall(pattern, text)) for pattern in chinese_patterns)
        
        # 가장 많이 매칭된 언어 반환
        language_scores = {
            'ko': korean_matches,
            'en': english_matches,
            'ja': japanese_matches,
            'zh': chinese_matches
        }
        
        detected_language = max(language_scores, key=language_scores.get)
        return detected_language

    def _is_similar_language_content(self, title, channel_title, reference_language):
        """
        현재 콘텐츠가 기준 언어와 유사한 언어인지 엄격하게 판단합니다.
        (개선된 로직) 원본 영상과 동일한 언어의 콘텐츠인지 '엄격하게' 판단합니다.
        """
        # 현재 콘텐츠의 언어를 감지합니다.
        content_language = self._detect_content_language(title, channel_title)
        
        # 원본 영상의 언어와 현재 콘텐츠의 언어가 정확히 일치하는 경우에만 True를 반환합니다.
        if content_language == reference_language:
            return True
        
        # 그 외 모든 경우는 관련 없는 언어로 판단하고 False를 반환합니다.
        return False

class LanguageManager(QDialog):
    """언어 관리자 다이얼로그"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Language Settings")  # 영어로 고정
        self.setModal(True)
        # 반응형 다이얼로그 설정 (기본 크기 400x300, 최소 크기 동일)
        self.resize(400, 300)
        self.setMinimumSize(400, 300)
        
        # 지원 언어 목록 (유튜브 주요 사용 국가들 포함)
        self.supported_languages = {
            'ko': '한국어',
            'en': 'English',
            'ja': '日本語',
            'zh': '中文',
            'es': 'Español',
            'fr': 'Français',
            'de': 'Deutsch',
            'ru': 'Русский',
            'pt': 'Português',
            'it': 'Italiano',
            'hi': 'हिन्दी (Hindi)',
            'ar': 'العربية (Arabic)',
            'tr': 'Türkçe',
            'th': 'ไทย (Thai)',
            'id': 'Bahasa Indonesia',
            'vi': 'Tiếng Việt',
            'nl': 'Nederlands',
            'pl': 'Polski',
            'sv': 'Svenska',
            'da': 'Dansk',
            'no': 'Norsk',
            'fi': 'Suomi',
            'uk': 'Українська'
        }
        
        self.current_language = 'ko'  # 기본 언어
        self.init_ui()
        self.load_language_settings()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # 안내 문구 (영어로 고정)
        info_label = QLabel("Please select your language:")
        info_label.setStyleSheet("font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(info_label)
        
        # 언어 목록
        self.language_list = QListWidget()
        self.language_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        
        # 언어 항목 추가
        for code, name in self.supported_languages.items():
            item = QListWidgetItem(name)
            item.setData(Qt.ItemDataRole.UserRole, code)
            self.language_list.addItem(item)
        
        layout.addWidget(self.language_list)
        
        # 안내 정보 (영어로 고정)
        notice_label = QLabel("※ Language will be applied after restarting the application.")
        notice_label.setStyleSheet("color: #666666; font-size: 10px; margin-top: 10px;")
        notice_label.setWordWrap(True)
        layout.addWidget(notice_label)
        
        # 버튼 (영어로 고정)
        button_layout = QHBoxLayout()
        
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept_language)
        self.ok_button.setStyleSheet("font-weight: bold;")
        
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        
        button_layout.addStretch()
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
    def load_language_settings(self):
        """언어 설정을 로드합니다."""
        try:
            with open(path_manager.CONFIG_PATH, 'r', encoding='utf-8') as f:
                config = json.load(f)
                self.current_language = config.get('language', 'ko')
        except FileNotFoundError:
            self.current_language = 'ko'
        except Exception as e:
            print(f"언어 설정 로드 오류: {e}")
            self.current_language = 'ko'
        
        # 현재 언어 선택
        for i in range(self.language_list.count()):
            item = self.language_list.item(i)
            if item.data(Qt.ItemDataRole.UserRole) == self.current_language:
                self.language_list.setCurrentItem(item)
                break
                
    def save_language_settings(self):
        """언어 설정을 저장합니다."""
        try:
            config = {}
            if os.path.exists(path_manager.CONFIG_PATH):
                with open(path_manager.CONFIG_PATH, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            
            config['language'] = self.current_language
            
            with open(path_manager.CONFIG_PATH, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"언어 설정 저장 오류: {e}")
            
    def accept_language(self):
        """언어 선택을 확인합니다."""
        current_item = self.language_list.currentItem()
        if current_item:
            self.current_language = current_item.data(Qt.ItemDataRole.UserRole)
            self.save_language_settings()
            
            # 언어 변경 시 즉시 적용
            setup_i18n(self.current_language)
            
            self.accept()
        else:
            QMessageBox.warning(self, "Warning", "Please select a language.")
            
    def get_selected_language(self):
        """선택된 언어 코드를 반환합니다."""
        return self.current_language

class APIKeyManager(QDialog):
    """API Key 관리 다이얼로그"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(_("저장된 API Key 관리"))
        # 반응형 다이얼로그 설정 (기본 크기 500x400, 최소 크기 동일)
        self.resize(500, 400)
        self.setMinimumSize(500, 400)
        self.setModal(True)
        
        self.api_keys_file = "api_keys.json"
        self.api_keys = []
        
        self.init_ui()
        self.load_api_keys()
        
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 제목
        title_label = QLabel(_("저장된 API Keys"))
        title_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addWidget(title_label)
        
        # API Key 목록
        self.api_key_list = QListWidget()
        self.api_key_list.setStyleSheet("""
            QListWidget {
                border: 1px solid #CCCCCC;
                border-radius: 3px;
                background-color: #FFFFFF;
                padding: 5px;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #EEEEEE;
                border-radius: 2px;
                margin: 1px 0px;
            }
            QListWidget::item:selected {
                background-color: #F0F0F0;
                border: 1px solid #0078D4;
                color: #333333;
            }
            QListWidget::item:hover {
                background-color: #F8F8F8;
            }
        """)
        layout.addWidget(self.api_key_list)
        
        # 새 API Key 입력
        input_layout = QHBoxLayout()
        self.new_key_input = QLineEdit()
        self.new_key_input.setPlaceholderText(_("새 API Key를 입력하세요..."))
        self.new_key_input.setFixedHeight(30)
        
        self.add_button = QPushButton(_("추가"))
        self.add_button.setFixedHeight(30)
        self.add_button.setFixedWidth(60)
        self.add_button.clicked.connect(self.add_api_key)
        
        input_layout.addWidget(self.new_key_input)
        input_layout.addWidget(self.add_button)
        layout.addLayout(input_layout)
        
        # 버튼들
        button_layout = QHBoxLayout()
        
        self.copy_button = QPushButton(_("복사"))
        self.copy_button.setFixedHeight(30)
        self.copy_button.clicked.connect(self.copy_selected_key)
        
        self.delete_button = QPushButton(_("삭제"))
        self.delete_button.setFixedHeight(30)
        self.delete_button.clicked.connect(self.delete_selected_key)
        
        self.use_button = QPushButton(_("사용"))
        self.use_button.setFixedHeight(30)
        self.use_button.clicked.connect(self.use_selected_key)
        
        self.close_button = QPushButton(_("닫기"))
        self.close_button.setFixedHeight(30)
        self.close_button.clicked.connect(self.close)
        
        button_layout.addWidget(self.copy_button)
        button_layout.addWidget(self.delete_button)
        button_layout.addWidget(self.use_button)
        button_layout.addStretch()
        button_layout.addWidget(self.close_button)
        
        layout.addLayout(button_layout)
        
        # 안내 문구
        info_label = QLabel(_("• 선택한 API Key를 복사하거나 삭제할 수 있습니다.\n• '사용' 버튼을 누르면 메인 화면의 API Key 입력란에 자동으로 입력됩니다.\n• 저장된 Key는 운영진에게 노출되지 않습니다. 외부 노출을 주의하십시오."))
        info_label.setStyleSheet("color: #666666; font-size: 10px; margin-top: 10px;")
        layout.addWidget(info_label)
        
    def load_api_keys(self):
        """저장된 API Key 목록을 로드합니다."""
        try:
            if os.path.exists(self.api_keys_file):
                with open(self.api_keys_file, 'r', encoding='utf-8') as f:
                    self.api_keys = json.load(f)
            else:
                self.api_keys = []
        except Exception as e:
            print(f"API Key 로드 오류: {e}")
            self.api_keys = []
        
        self.update_list()
    
    def save_api_keys(self):
        """API Key 목록을 파일에 저장합니다."""
        try:
            with open(self.api_keys_file, 'w', encoding='utf-8') as f:
                json.dump(self.api_keys, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"API Key 저장 오류: {e}")
    
    def update_list(self):
        """목록을 업데이트합니다."""
        self.api_key_list.clear()
        for i, key in enumerate(self.api_keys):
            # API Key를 마스킹해서 표시 (앞 10자리만 표시)
            masked_key = key[:10] + "..." + key[-5:] if len(key) > 15 else key
            item = QListWidgetItem(f"{i+1}. {masked_key}")
            item.setData(Qt.ItemDataRole.UserRole, key)  # 원본 키 저장
            self.api_key_list.addItem(item)
    
    def add_api_key(self):
        """새 API Key를 추가합니다."""
        new_key = self.new_key_input.text().strip()
        if not new_key:
            QMessageBox.warning(self, _("경고"), _("API Key를 입력해주세요."))
            return
        
        if new_key in self.api_keys:
            QMessageBox.warning(self, _("경고"), _("이미 존재하는 API Key입니다."))
            return
        
        self.api_keys.append(new_key)
        self.save_api_keys()
        self.update_list()
        self.new_key_input.clear()
        
        QMessageBox.information(self, _("성공"), _("API Key가 추가되었습니다."))
    
    def copy_selected_key(self):
        """선택된 API Key를 복사합니다."""
        current_item = self.api_key_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, _("경고"), _("복사할 API Key를 선택해주세요."))
            return
        
        api_key = current_item.data(Qt.ItemDataRole.UserRole)
        clipboard = QApplication.clipboard()
        clipboard.setText(api_key)
        
        QMessageBox.information(self, _("성공"), _("API Key가 클립보드에 복사되었습니다."))
    
    def delete_selected_key(self):
        """선택된 API Key를 삭제합니다."""
        current_item = self.api_key_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, _("경고"), _("삭제할 API Key를 선택해주세요."))
            return
        
        reply = QMessageBox.question(
            self, "확인", 
            "선택한 API Key를 삭제하시겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            api_key = current_item.data(Qt.ItemDataRole.UserRole)
            self.api_keys.remove(api_key)
            self.save_api_keys()
            self.update_list()
            
            QMessageBox.information(self, _("성공"), _("API Key가 삭제되었습니다."))
    
    def use_selected_key(self):
        """선택된 API Key를 사용합니다."""
        current_item = self.api_key_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, _("경고"), _("사용할 API Key를 선택해주세요."))
            return
        
        self.selected_key = current_item.data(Qt.ItemDataRole.UserRole)
        self.accept()  # 다이얼로그 닫기
    
    def get_selected_key(self):
        """선택된 API Key를 반환합니다."""
        return getattr(self, 'selected_key', None)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        
        self.setWindowTitle(_("유튜브 딥서치"))
        # 반응형 창 크기 설정 (기본 크기 880x750, 최소 크기도 동일)
        self.resize(880, 750)
        self.setMinimumSize(880, 750)
        self.setStyleSheet(STYLESHEET)
        self.worker = None
        
        # 페이지네이션 상태 변수
        self.all_results = []
        self.current_page = 1
        self.results_per_page = 30
        self.total_pages = 0
        
        # 경로 설정 변수
        self.excel_save_path = None
        self.thumbnail_save_path = None
        self.settings_file = "app_settings.json"
        
        # 제목 확장 기능을 위한 변수
        self.expanded_videos = set()  # 확장된 비디오 ID 저장
        
        # Shift 키 범위 선택을 위한 변수
        self.last_selected_row = -1
        
        # 전역 선택 상태 관리 (전체 결과에서의 인덱스 기준)
        self.global_selected_items = set()  # 선택된 전체 결과의 인덱스들
        
        # 언어 번역 시스템 초기화
        self.translator = QTranslator()
        self.current_language = 'ko'
        
        # 설정 로드
        self.load_settings()
        
        self.init_ui()
        self.update_language_button()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(15, 10, 15, 15)  # 반응형을 위해 여백 조정
        main_layout.setSpacing(8)  # 반응형을 위해 간격 조정
        
        # === 검색 설정 영역 ===
        settings_widget = QWidget()
        settings_main_layout = QHBoxLayout(settings_widget)
        settings_main_layout.setSpacing(20)  # 반응형을 위해 spacing 조정
        settings_main_layout.setContentsMargins(10, 5, 10, 5)  # 상하 여백 줄임

        # 좌측 컬럼 생성
        left_column = QWidget()
        left_layout = QGridLayout(left_column)
        left_layout.setVerticalSpacing(8)  # 행 간의 수직 간격
        left_layout.setHorizontalSpacing(12)  # 열 간의 수평 간격
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setColumnMinimumWidth(0, 120)
        left_layout.setColumnStretch(1, 1)
        
        # 행 높이 통일을 위한 설정 - 완전히 정렬된 레이아웃
        row_height = 40  # 모든 행이 동일한 높이를 가지도록 설정
        left_layout.setRowMinimumHeight(0, row_height)
        left_layout.setRowMinimumHeight(1, row_height)
        left_layout.setRowMinimumHeight(2, row_height)
        left_layout.setRowMinimumHeight(3, row_height)
        left_layout.setRowMinimumHeight(4, row_height)
        
        # 좌측: API Key 라벨
        api_key_label = QLabel(_("API Key"))
        api_key_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # type: ignore # 텍스트를 입력필드 정중앙에 맞춤
        left_layout.addWidget(api_key_label, 0, 0)
        
        # 좌측: API Key 입력필드
        self.api_key_input = QLineEdit()
        self.api_key_input.setText(self.last_used_api_key)
        self.api_key_input.setEchoMode(QLineEdit.Password)  # type: ignore
        self.api_key_input.setEnabled(True)
        self.api_key_input.setFixedHeight(30)  # 통일된 높이
        self.api_key_input.setMinimumWidth(150)  # 반응형을 위해 최소 너비 조정
        left_layout.addWidget(self.api_key_input, 0, 1, Qt.AlignVCenter)  # type: ignore # 입력필드를 라벨과 동일한 높이에 맞춤
        
        # 좌측: 검색 키워드 또는 영상 링크
        keyword_label = QLabel(_("키워드/영상 url"))
        keyword_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # type: ignore
        left_layout.addWidget(keyword_label, 1, 0)
        
        self.keyword_input = QLineEdit()
        self.keyword_input.setText(_("당뇨"))
        self.keyword_input.setFixedHeight(30)  # 통일된 높이
        self.keyword_input.setMinimumWidth(150)  # 반응형을 위해 최소 너비 조정
        left_layout.addWidget(self.keyword_input, 1, 1, Qt.AlignVCenter)  # type: ignore
        
        # 좌측: 최소 조회수
        min_views_label = QLabel(_("최소 조회수"))
        min_views_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # type: ignore
        left_layout.addWidget(min_views_label, 2, 0)
        
        self.min_views_input = QComboBox()
        # userData 방식으로 데이터와 UI 분리
        min_views_data = {
            '0': _('제한 없음'),
            '10000': _('1만'),
            '30000': _('3만'),
            '50000': _('5만'),
            '100000': _('10만'),
            '150000': _('15만'),
            '200000': _('20만'),
            '500000': _('50만'),
            '1000000': _('100만')
        }
        for value, text in min_views_data.items():
            self.min_views_input.addItem(text, userData=value)
        self.min_views_input.setFixedHeight(30)  # 통일된 높이
        self.min_views_input.setMinimumWidth(150)  # 반응형을 위해 최소 너비 조정
        left_layout.addWidget(self.min_views_input, 2, 1, Qt.AlignVCenter)  # type: ignore
        
        # 좌측: 업로드 기간
        upload_period_label = QLabel(_("업로드 기간"))
        upload_period_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # type: ignore
        left_layout.addWidget(upload_period_label, 3, 0)
        
        self.upload_period_input = QComboBox()
        # userData 방식으로 데이터와 UI 분리
        upload_period_data = {
            'all': _('전체 기간'),
            'day': _('1일'),
            'week': _('1주일'),
            'month': _('1개월'),
            'month2': _('2개월'),
            'month3': _('3개월'),
            'month6': _('6개월'),
            'year': _('1년')
        }
        for value, text in upload_period_data.items():
            self.upload_period_input.addItem(text, userData=value)
        self.upload_period_input.setCurrentIndex(4)  # '2개월' 인덱스
        self.upload_period_input.setFixedHeight(30)  # 통일된 높이
        self.upload_period_input.setMinimumWidth(150)  # 반응형을 위해 최소 너비 조정
        left_layout.addWidget(self.upload_period_input, 3, 1, Qt.AlignVCenter)  # type: ignore
        
        # 좌측: 동영상 길이 (위치 조정)
        duration_label = QLabel(_("동영상 길이"))
        duration_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # type: ignore
        left_layout.addWidget(duration_label, 4, 0)
        
        duration_widget = QWidget()
        duration_layout = QHBoxLayout(duration_widget)
        duration_layout.setContentsMargins(0, 0, 0, 0)  # 좌측 여백을 0으로 설정
        duration_layout.setSpacing(15)
        
        self.duration_any_radio = QRadioButton(_("전체"))
        self.duration_short_radio = QRadioButton(_("쇼츠"))
        self.duration_long_radio = QRadioButton(_("롱폼"))
        self.duration_long_radio.setChecked(True)
        
        # 라디오 버튼의 스타일을 직접 설정하여 잘림 방지
        for radio in [self.duration_any_radio, self.duration_short_radio, self.duration_long_radio]:
            radio.setStyleSheet("""
                QRadioButton {
                    spacing: 8px;
                    margin-left: 2px;
                }
                QRadioButton::indicator {
                    width: 14px;
                    height: 14px;
                    margin-right: 5px;
                }
            """)
        
        duration_layout.addWidget(self.duration_any_radio)
        duration_layout.addWidget(self.duration_short_radio)
        duration_layout.addWidget(self.duration_long_radio)
        duration_layout.addStretch()
        
        left_layout.addWidget(duration_widget, 4, 1, Qt.AlignVCenter)  # type: ignore
        
        # 우측 컬럼 생성
        right_column = QWidget()
        right_layout = QGridLayout(right_column)
        right_layout.setVerticalSpacing(8)  # 좌측과 동일한 수직 간격
        right_layout.setHorizontalSpacing(12)  # 좌측과 동일한 수평 간격
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setColumnMinimumWidth(0, 120)
        right_layout.setColumnStretch(1, 1)
        
        # 행 높이 통일을 위한 설정 (좌측과 완전히 동일) - 완전히 정렬된 레이아웃
        right_layout.setRowMinimumHeight(0, row_height)
        right_layout.setRowMinimumHeight(1, row_height)
        right_layout.setRowMinimumHeight(2, row_height)
        right_layout.setRowMinimumHeight(3, row_height)
        right_layout.setRowMinimumHeight(4, row_height)

        # 우측: 저장된 API key 링크와 언어 설정
        api_language_layout = QHBoxLayout()
        
        # 저장된 API key 링크 스타일 텍스트
        self.saved_api_key_label = QLabel(_("저장된 API key"))
        self.saved_api_key_label.setStyleSheet("""
            QLabel {
                color: #0078D4;
                text-decoration: underline;
                background-color: transparent;
                font-size: 12px;
            }
            QLabel:hover {
                color: #106EBE;
                cursor: pointer;
            }
        """)
        self.saved_api_key_label.mousePressEvent = lambda event: self.open_api_key_manager()
        
        # 언어 설정 버튼
        self.language_button = QPushButton("Language")
        self.language_button.setFixedSize(80, 24)
        self.language_button.setStyleSheet("""
            QPushButton {
                background-color: #F5F5F5;
                border: 1px solid #CCCCCC;
                border-radius: 3px;
                padding: 2px 8px;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #E8E8E8;
            }
            QPushButton:pressed {
                background-color: #D0D0D0;
            }
        """)
        self.language_button.clicked.connect(self.open_language_manager)
        
        api_language_layout.addWidget(self.saved_api_key_label)
        api_language_layout.addStretch()
        api_language_layout.addWidget(self.language_button)
        
        right_layout.addLayout(api_language_layout, 0, 0, 1, 2, Qt.AlignVCenter)  # type: ignore  # 2개 컬럼에 걸쳐 배치
        
        # 우측: 최대 구독자 수 (좌측 검색 키워드와 같은 행)
        max_subs_label = QLabel(_("최대 구독자 수"))
        max_subs_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # type: ignore
        max_subs_label.setFixedHeight(30)  # 좌측 라벨과 동일한 높이
        right_layout.addWidget(max_subs_label, 1, 0, Qt.AlignVCenter)  # type: ignore
        
        self.max_subs_input = QSpinBox()
        self.max_subs_input.setButtonSymbols(QSpinBox.NoButtons)  # type: ignore
        self.max_subs_input.setRange(0, 1000000000)  # 10억까지 (현실적인 최대값)
        self.max_subs_input.setSingleStep(1000)
        self.max_subs_input.setValue(10000)
        self.max_subs_input.setFixedHeight(30)  # 좌측 입력 필드와 동일한 높이
        self.max_subs_input.setMinimumWidth(150)  # 반응형을 위해 최소 너비 조정
        # 숫자만 입력 가능하도록 설정 (QSpinBox는 기본적으로 숫자만 허용하지만 명시적으로 설정)
        self.max_subs_input.setKeyboardTracking(False)  # 타이핑 중간 값 변경 방지
        right_layout.addWidget(self.max_subs_input, 1, 1, Qt.AlignVCenter)  # type: ignore
        
        # 우측: 정렬 방식 (좌측 최소 조회수와 같은 행)
        sort_by_label = QLabel(_("정렬 방식"))
        sort_by_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # type: ignore
        sort_by_label.setFixedHeight(30)  # 좌측 라벨과 동일한 높이
        right_layout.addWidget(sort_by_label, 2, 0, Qt.AlignVCenter)  # type: ignore
        
        self.sort_by_input = QComboBox()
        # userData 방식으로 데이터와 UI 분리
        sort_by_data = {
            'viewCount': _('조회수'),
            'engagement_rate': _('참여도 (좋아요 / 조회수)'),
            'reaction_rate': _('반응도 (댓글 / 조회수)'),
            'date_desc': _('날짜 ↑ (최신순)'),
            'date_asc': _('날짜 ↓ (오래된순)')
        }
        for value, text in sort_by_data.items():
            self.sort_by_input.addItem(text, userData=value)
        self.sort_by_input.setFixedHeight(30)  # 좌측 입력 필드와 동일한 높이
        self.sort_by_input.setMinimumWidth(150)  # 반응형을 위해 최소 너비 조정
        right_layout.addWidget(self.sort_by_input, 2, 1, Qt.AlignVCenter)  # type: ignore
        
        # 우측: 저장 경로 설정 (좌측 업로드 기간과 같은 행)
        path_setting_label = QLabel(_("저장 경로 설정"))
        # AlignVCenter를 사용하여 좌측 업로드 기간과 동일한 선상에 배치
        path_setting_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # type: ignore
        path_setting_label.setFixedHeight(30)  # 좌측 라벨과 동일한 높이
        right_layout.addWidget(path_setting_label, 3, 0, Qt.AlignVCenter) # type: ignore

        # 버튼들을 라벨과 같은 행에 배치하기 위한 위젯 생성
        buttons_container = QWidget()
        buttons_container.setFixedHeight(30)  # 좌측 입력 필드와 동일한 높이
        buttons_layout = QHBoxLayout(buttons_container)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(8)
        
        self.excel_path_button = QPushButton(_("엑셀 저장 경로"))
        self.excel_path_button.setFixedHeight(30)  # 좌측 입력 필드와 동일한 높이
        self.excel_path_button.setMinimumWidth(70)  # 반응형을 위해 최소 너비 조정
        self.excel_path_button.clicked.connect(self.select_excel_path)
        
        self.thumbnail_path_button = QPushButton(_("썸네일 저장 경로"))
        self.thumbnail_path_button.setFixedHeight(30)  # 좌측 입력 필드와 동일한 높이
        self.thumbnail_path_button.setMinimumWidth(70)  # 반응형을 위해 최소 너비 조정
        self.thumbnail_path_button.clicked.connect(self.select_thumbnail_path)
        
        # 버튼들이 전체 너비를 1:1로 나누어 가지도록 설정
        buttons_layout.addWidget(self.excel_path_button, 1)
        buttons_layout.addWidget(self.thumbnail_path_button, 1)
        
        # 버튼 컨테이너를 라벨과 같은 행에 배치
        right_layout.addWidget(buttons_container, 3, 1, Qt.AlignVCenter)  # type: ignore
        
        # 경로 표시 라벨들을 새로운 행에 배치
        path_display_widget = QWidget()
        path_display_layout = QVBoxLayout(path_display_widget)
        path_display_layout.setContentsMargins(0, 0, 0, 0)
        path_display_layout.setSpacing(2)
        
        self.excel_path_label = QLabel(_("엑셀: 다운로드 폴더"))
        self.excel_path_label.setStyleSheet("color: #666666; font-size: 9px;")
        
        self.thumbnail_path_label = QLabel(_("썸네일: 다운로드 폴더"))
        self.thumbnail_path_label.setStyleSheet("color: #666666; font-size: 9px;")
        
        # 안내 문구
        info_label = QLabel(_("*경로 미지정 시, 다운로드 폴더에 저장됩니다."))
        info_label.setStyleSheet("color: #888888; font-size: 9px;")
        
        path_display_layout.addWidget(self.excel_path_label)
        path_display_layout.addWidget(self.thumbnail_path_label)
        path_display_layout.addWidget(info_label)
        
        # 경로 표시 위젯을 새로운 행에 배치 (4번째 행)
        right_layout.addWidget(path_display_widget, 4, 1, Qt.AlignTop)  # type: ignore
        
        # 메인 레이아웃에 컬럼들 추가 (반응형을 위해 stretch factor 설정)
        settings_main_layout.addWidget(left_column, 1)   # 좌측 컬럼 stretch factor 1
        settings_main_layout.addWidget(right_column, 1)  # 우측 컬럼 stretch factor 1
        
        main_layout.addWidget(settings_widget)
        
        # === 검색 버튼 영역 ===
        search_layout = QHBoxLayout()
        self.search_button = QPushButton(_("검색"))
        self.search_button.clicked.connect(self.start_search)
        self.search_button.setFixedWidth(100)
        
        search_layout.addWidget(self.search_button)
        search_layout.addStretch()
        main_layout.addLayout(search_layout)
        
        # === 페이지네이션 영역 ===
        pagination_layout = QHBoxLayout()
        
        self.prev_page_button = QPushButton(_("이전 페이지"))
        self.prev_page_button.setEnabled(False)
        self.prev_page_button.setFixedWidth(100)
        self.prev_page_button.clicked.connect(self.prev_page)
        
        self.page_info_label = QLabel("")
        self.page_info_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # type: ignore
        self.page_info_label.setStyleSheet("font-weight: bold; background-color: transparent; margin-left: 10px; margin-right: 10px;")
        
        self.next_page_button = QPushButton(_("다음 페이지"))
        self.next_page_button.setEnabled(False)
        self.next_page_button.setFixedWidth(100)
        self.next_page_button.clicked.connect(self.next_page)
        
        pagination_layout.addWidget(self.prev_page_button)
        pagination_layout.addWidget(self.page_info_label)
        pagination_layout.addWidget(self.next_page_button)
        pagination_layout.addStretch()  # 우측에 공간 확보
        
        main_layout.addLayout(pagination_layout)
        
        # === 결과 리스트 영역 ===
        # 결과 리스트 라벨과 체크박스를 같은 선상에 배치
        results_header_layout = QHBoxLayout()
        
        self.results_label = QLabel(_("결과 리스트"))
        self.results_label.setStyleSheet("font-weight: bold; background-color: transparent;")
        results_header_layout.addWidget(self.results_label)
        
        results_header_layout.addStretch()  # 중간 공간 확보
        
        # CC 필터 체크박스 추가 (결과 리스트와 같은 선상)
        self.cc_filter_checkbox = QCheckBox(_("재사용 가능 콘텐츠만 보기 (크리에이터 커먼즈)"))
        self.cc_filter_checkbox.setStyleSheet("""
            QCheckBox {
                font-size: 11px;
                color: #666666;
                spacing: 8px;
            }
            QCheckBox:hover {
                color: #0078D4;
            }
            QCheckBox:disabled {
                color: #CCCCCC;
            }
        """)
        self.cc_filter_checkbox.setEnabled(False)  # 초기에는 비활성화
        self.cc_filter_checkbox.stateChanged.connect(self.filter_cc_videos)
        results_header_layout.addWidget(self.cc_filter_checkbox)
        
        main_layout.addLayout(results_header_layout)
        
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(8)
        self.table_widget.setHorizontalHeaderLabels([_("선택"), _("업로드 날짜"), _("조회수"), _("분량"), _("제목"), _("채널명"), _("구독자 수"), "Data"])
        self.table_widget.setColumnHidden(7, True)  # 데이터 저장용 숨김 컬럼
        
        # 테이블 컬럼 크기 설정 (반응형으로 변경)
        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)  # type: ignore # 선택 - 고정
        header.setSectionResizeMode(1, QHeaderView.Fixed)  # type: ignore # 업로드 날짜 - 고정 (일관성을 위해)
        header.setSectionResizeMode(2, QHeaderView.Fixed)  # type: ignore # 조회수 - 고정 (일관성을 위해)
        header.setSectionResizeMode(3, QHeaderView.Fixed)  # type: ignore # 분량 - 고정 너비로 변경
        header.setSectionResizeMode(4, QHeaderView.Stretch)  # type: ignore # 제목 - 남은 공간 모두 활용
        header.setSectionResizeMode(5, QHeaderView.Fixed)  # type: ignore # 채널명 - 고정 (일관성을 위해)
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # type: ignore # 구독자 수 - 내용에 맞춤 (우선 표시)
        
        # 반응형을 위한 컬럼 크기 최적화
        header.resizeSection(0, 25)   # 선택 - 25px (체크박스)
        header.resizeSection(1, 70)   # 업로드 날짜 - 70px (반응형 조정)
        header.resizeSection(2, 55)   # 조회수 - 55px (반응형 조정)
        header.resizeSection(3, 30)   # 분량 - 30px 
        header.resizeSection(5, 90)   # 채널명 - 90px (반응형 조정)
        
        # 구독자 수 컬럼의 최소 너비 설정 (반응형 조정)
        header.setMinimumSectionSize(45)  # 구독자 수 최소 너비 (반응형)
        
        # 반응형에서는 필요시 가로 스크롤바 표시
        self.table_widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # type: ignore
        
        # 순번(행 헤더) 중앙 정렬
        self.table_widget.verticalHeader().setDefaultAlignment(Qt.AlignCenter) # type: ignore

        # 순번(행 헤더) 중앙 정렬을 위한 스타일시트 정리
        self.table_widget.setStyleSheet("""
            QTableWidget {
                border: 1px solid #CCCCCC;
                gridline-color: #E0E0E0;
                background-color: #FFFFFF;
            }
            QHeaderView::section:vertical {
                background-color: #F0F0F0;
                padding: 4px;
                border: 1px solid #CCCCCC;
                font-weight: bold;
                font-size: 12px;
            }
        """)
        
        self.table_widget.setEditTriggers(QAbstractItemView.NoEditTriggers)  # type: ignore
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)  # type: ignore
        
        # 제목 클릭 이벤트 연결
        self.table_widget.cellClicked.connect(self.on_title_clicked)
        
        main_layout.addWidget(self.table_widget)

        # === 하단 버튼 영역 ===
        bottom_layout = QHBoxLayout()
        
        self.select_all_button = QPushButton(_("모두 선택"))
        self.deselect_all_button = QPushButton(_("모두 해제"))
        
        # 선택 개수 표시 라벨
        self.selected_count_label = QLabel(_("선택됨: 0개"))
        self.selected_count_label.setStyleSheet("font-weight: bold; background-color: transparent; color: #0078D4;")
        self.selected_count_label.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)  # type: ignore
        self.excel_button = QPushButton(_("엑셀 추출"))
        self.channel_link_button = QPushButton(_("채널 바로가기"))
        self.video_link_button = QPushButton(_("영상 바로가기"))
        self.thumbnail_button = QPushButton(_("썸네일 추출"))

        # 버튼 크기 통일
        for btn in [self.select_all_button, self.deselect_all_button, self.excel_button, 
                   self.channel_link_button, self.video_link_button, self.thumbnail_button]:
            btn.setFixedHeight(30)

        # 이벤트 연결
        self.select_all_button.clicked.connect(lambda: self.toggle_select_all(True))
        self.deselect_all_button.clicked.connect(lambda: self.toggle_select_all(False))
        self.excel_button.clicked.connect(self.export_to_excel)
        self.channel_link_button.clicked.connect(self.open_channel_link)
        self.video_link_button.clicked.connect(self.open_video_link)
        self.thumbnail_button.clicked.connect(self.download_thumbnails)
        
        bottom_layout.addWidget(self.select_all_button)
        bottom_layout.addWidget(self.deselect_all_button)
        bottom_layout.addWidget(self.selected_count_label)
        bottom_layout.addStretch()  # 가변 여백으로 반응형 지원
        bottom_layout.addWidget(self.channel_link_button)
        bottom_layout.addWidget(self.video_link_button)
        bottom_layout.addWidget(self.excel_button)
        bottom_layout.addWidget(self.thumbnail_button)
        main_layout.addLayout(bottom_layout)
        
        # 경로 표시 초기화
        self.update_path_display()
    
    def open_api_key_manager(self):
        """API Key 관리 다이얼로그를 엽니다."""
        dialog = APIKeyManager(self)
        if dialog.exec() == QDialog.Accepted:  # type: ignore
            selected_key = dialog.get_selected_key()
            if selected_key:
                self.api_key_input.setText(selected_key)
                QMessageBox.information(self, _("성공"), _("API Key가 적용되었습니다."))

    def on_title_clicked(self, row, column):
        """제목 컬럼 클릭시 제목 확장/축소"""
        if column == 4:  # 제목 컬럼인 경우
            title_item = self.table_widget.item(row, column)
            data_item = self.table_widget.item(row, 7)  # 숨김 데이터 컬럼
            
            if title_item and data_item:
                full_title = title_item.text()
                video_id = data_item.text().split('|')[0]  # video_id 추출
                
                # 기존 확장 행들을 모두 제거
                self._remove_all_expanded_rows()
                
                if video_id in self.expanded_videos:
                    # 확장된 상태 -> 축소
                    self.expanded_videos.remove(video_id)
                else:
                    # 축소된 상태 -> 확장
                    self.expanded_videos.clear()  # 다른 확장 상태 모두 해제
                    self.expanded_videos.add(video_id)
                    
                    # 확장 행 제거 후 정확한 행 위치 재계산
                    correct_row = self._find_video_row(video_id)
                    if correct_row is not None:
                        # 새 행 삽입
                        self.table_widget.insertRow(correct_row + 1)
                        
                        # 현재 페이지에서의 글로벌 인덱스 계산
                        start_idx = (self.current_page - 1) * self.results_per_page
                        global_idx = start_idx + correct_row
                        
                        # 확장 UI 구성 - QWidget + QVBoxLayout 사용 (세로 배치, 여유 있는 높이)
                        expanded_widget = QWidget()
                        expanded_widget.setMaximumHeight(80)  # 최대 높이 증가 (60→80)
                        layout = QVBoxLayout(expanded_widget)
                        layout.setContentsMargins(8, 6, 8, 6)  # 여백 약간 증가
                        layout.setSpacing(6)  # 간격 약간 증가
                        
                        # 전체 제목 QLabel - 위쪽에 완전 표시 (충분한 높이 확보)
                        title_label = QLabel(full_title)
                        title_label.setWordWrap(True)
                        title_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
                        title_label.setMaximumHeight(60)  # 최대 높이 증가 (40→60)
                        title_label.setStyleSheet("""
                            font-size: 11px; 
                            color: #333333; 
                            background: transparent; 
                            border: none; 
                            padding: 2px;
                            margin-bottom: 3px;
                            font-weight: bold;
                            line-height: 1.2;
                        """)
                        layout.addWidget(title_label)
                        
                        # 기여도 관련 정보가 이미 있는지 확인
                        contribution_info = None
                        if global_idx < len(self.all_results):
                            item_data = self.all_results[global_idx]
                            if 'contribution_score' in item_data and 'channel_avg_views' in item_data:
                                contribution_score = item_data['contribution_score']
                                channel_avg_views = item_data['channel_avg_views']
                                contribution_info = _("평균 대비 {:.1f}% 성과 (채널 평균: {:,}회)").format(contribution_score, channel_avg_views)
                        
                        # 기여도 관련 위젯들을 담을 컨테이너
                        contribution_container = QWidget()
                        contribution_layout = QHBoxLayout(contribution_container)
                        contribution_layout.setContentsMargins(0, 0, 0, 0)
                        contribution_layout.setAlignment(Qt.AlignLeft)
                        
                        if contribution_info:
                            # 기여도 정보가 이미 있으면 바로 표시
                            result_label = QLabel(contribution_info)
                            result_label.setStyleSheet("""
                                font-size: 10px; 
                                color: #666666; 
                                font-weight: bold;
                                background: #E8F5E8;
                                padding: 3px 6px;
                                border-radius: 2px;
                                max-width: 300px;
                            """)
                            contribution_layout.addWidget(result_label)
                        else:
                            # 기여도 정보가 없으면 "기여도 확인하기" 버튼 표시
                            contrib_button = QPushButton(_("기여도(성과) 확인하기"))
                            contrib_button.setObjectName(f"contrib_btn_{correct_row}")
                            contrib_button.setStyleSheet("""
                                QPushButton {
                                    background-color: #4CAF50;
                                    color: white;
                                    border: none;
                                    padding: 4px 8px;
                                    border-radius: 3px;
                                    font-size: 10px;
                                    font-weight: bold;
                                    max-width: 120px;
                                }
                                QPushButton:hover {
                                    background-color: #45a049;
                                }
                                QPushButton:disabled {
                                    background-color: #cccccc;
                                }
                            """)
                            
                            # 버튼 클릭 시 기여도 분석 요청
                            if global_idx < len(self.all_results):
                                item_data = self.all_results[global_idx]
                                channel_id = item_data.get('channel_id')
                                video_id = item_data.get('video_id')
                                view_count = item_data.get('view_count', 0)
                                
                                contrib_button.clicked.connect(
                                    lambda checked, btn=contrib_button, r=correct_row, ch=channel_id, 
                                    vi=video_id, vc=view_count, gi=global_idx: 
                                    self.request_contribution_data(btn, r, ch, vi, vc, gi)
                                )
                            
                            contribution_layout.addWidget(contrib_button)
                        
                        contribution_layout.addStretch()  # 우측에 여백 추가
                        layout.addWidget(contribution_container)
                        
                        # 전체 위젯 스타일링 - 깔끔하게 단순화
                        expanded_widget.setStyleSheet("""
                            QWidget {
                                background-color: #FAFAFA;
                                border: none;
                                border-left: 3px solid #4CAF50;
                                border-radius: 0px;
                            }
                        """)
                        
                        # 확장된 위젯을 전체 컬럼에 걸쳐 표시
                        self.table_widget.setCellWidget(correct_row + 1, 0, expanded_widget)
                        self.table_widget.setSpan(correct_row + 1, 0, 1, 8)  # 모든 컬럼에 걸쳐 병합
                        
                        # 확장 행의 높이 조정 (50→70으로 증가)
                        self.table_widget.setRowHeight(correct_row + 1, 70)
                        
                        # 확장 행의 헤더를 완전히 숨기기
                        empty_header = QTableWidgetItem("")
                        self.table_widget.setVerticalHeaderItem(correct_row + 1, empty_header)
    
    def _find_video_row(self, video_id):
        """비디오 ID로 정확한 행 위치를 찾습니다."""
        for row in range(self.table_widget.rowCount()):
            data_item = self.table_widget.item(row, 7)
            if data_item and data_item.text().split('|')[0] == video_id:
                return row
        return None
    
    def request_contribution_data(self, button, row, channel_id, video_id, view_count, global_idx):
        """기여도 분석 요청 - 주문형 분석 시작"""
        if not channel_id or not video_id:
            QMessageBox.warning(self, _("오류"), _("기여도 분석에 필요한 데이터가 부족합니다."))
            return
        
        # 현재 UI에 입력된 API 키 가져오기
        api_key = self.api_key_input.text().strip()
        if not api_key:
            QMessageBox.warning(self, _("오류"), _("API 키가 설정되지 않았습니다."))
            return
        
        # 버튼 상태 변경
        button.setText(_("분석 중..."))
        button.setEnabled(False)
        
        # ContributionWorker 생성 및 실행
        self.contribution_worker = ContributionWorker(
            api_key=api_key,
            channel_id=channel_id,
            video_id=video_id,
            view_count=view_count,
            row=global_idx  # 글로벌 인덱스 전달
        )
        
        # 시그널 연결
        self.contribution_worker.finished.connect(self.update_contribution_ui)
        self.contribution_worker.error.connect(self.on_contribution_error)
        
        # 워커 시작
        self.contribution_worker.start()
        
        # 워커 참조 저장 (가비지 컬렉션 방지)
        self.contribution_worker.button_ref = button
        self.contribution_worker.table_row = row
    
    def update_contribution_ui(self, global_idx, contribution_data):
        """기여도 분석 결과를 UI에 반영"""
        try:
            # all_results에 기여도 데이터 저장 (엑셀 추출용)
            if global_idx < len(self.all_results):
                self.all_results[global_idx]['contribution_score'] = contribution_data.get('contribution_score', 0)
                self.all_results[global_idx]['channel_avg_views'] = contribution_data.get('channel_avg_views', 0)
            
            # 현재 확장된 행 찾기 및 UI 업데이트 (새로운 QVBoxLayout 구조)
            for row in range(self.table_widget.rowCount()):
                expanded_widget = self.table_widget.cellWidget(row, 0)
                if expanded_widget and hasattr(expanded_widget, 'layout'):
                    main_layout = expanded_widget.layout()  # QVBoxLayout
                    if main_layout and main_layout.count() >= 2:
                        # contribution_container 찾기 (두 번째 위젯)
                        contribution_container = main_layout.itemAt(1).widget()
                        if contribution_container and hasattr(contribution_container, 'layout'):
                            contribution_layout = contribution_container.layout()  # QHBoxLayout
                            
                            # 기존 버튼 찾기
                            button_found = False
                            for i in range(contribution_layout.count()):
                                item = contribution_layout.itemAt(i)
                                if item and item.widget():
                                    w = item.widget()
                                    if isinstance(w, QPushButton) and w.objectName().startswith("contrib_btn_"):
                                        # 버튼을 결과 라벨로 교체
                                        contribution_score = contribution_data.get('contribution_score', 0)
                                        channel_avg_views = contribution_data.get('channel_avg_views', 0)
                                        result_text = _("평균 대비 {:.1f}% 성과 (채널 평균: {:,}회)").format(contribution_score, channel_avg_views)
                                        
                                        result_label = QLabel(result_text)
                                        result_label.setStyleSheet("""
                                            font-size: 10px; 
                                            color: #666666; 
                                            font-weight: bold;
                                            background: #E8F5E8;
                                            padding: 3px 6px;
                                            border-radius: 2px;
                                            max-width: 300px;
                                        """)
                                        
                                        # 버튼 제거하고 라벨 추가
                                        contribution_layout.removeWidget(w)
                                        w.deleteLater()
                                        contribution_layout.insertWidget(i, result_label)
                                        button_found = True
                                        break
                            
                            if button_found:
                                break
            
            # 워커 정리
            if hasattr(self, 'contribution_worker'):
                self.contribution_worker.deleteLater()
                delattr(self, 'contribution_worker')
        
        except Exception as e:
            print(f"UI 업데이트 오류: {e}")
    
    def on_contribution_error(self, error_message):
        """기여도 분석 오류 처리"""
        # 구체적인 오류 코드별 처리
        if error_message == "QUOTA_EXCEEDED":
            QMessageBox.critical(
                self, 
                _("API 할당량 초과"), 
                _("금일 API 할당량을 전부 소진했습니다.\n\n기여도 분석을 위해서는 내일 다시 시도하시거나 다른 API 키를 사용해주세요.")
            )
        elif error_message == "KEY_INVALID":
            QMessageBox.critical(
                self, 
                _("API 키 오류"), 
                _("올바르지 않은 API 키입니다.\n\nAPI 키를 다시 확인해주세요.")
            )
        else:
            QMessageBox.critical(self, _("기여도 분석 오류"), error_message)
        
        # 버튼 상태 복원
        try:
            if hasattr(self, 'contribution_worker') and hasattr(self.contribution_worker, 'button_ref'):
                button = self.contribution_worker.button_ref
                button.setText(_("기여도 확인하기"))
                button.setEnabled(True)
        except:
            pass
        
        # 워커 정리
        if hasattr(self, 'contribution_worker'):
            self.contribution_worker.deleteLater()
            delattr(self, 'contribution_worker')
    
    def get_current_api_key(self):
        """현재 설정된 API 키 반환"""
        try:
            with open(path_manager.SETTINGS_PATH, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                return settings.get('last_api_key', '')
        except:
            return ''
    
    def _remove_all_expanded_rows(self):
        """모든 확장 행을 제거합니다."""
        rows_to_remove = []
        for row in range(self.table_widget.rowCount()):
            # 확장 행의 정확한 특징을 확인
            widget = self.table_widget.cellWidget(row, 0)
            if widget is not None:
                # 확장 행은 0번 컬럼에 QLabel이 설정됨 (체크박스가 아님)
                from PySide6.QtWidgets import QLabel
                if isinstance(widget, QLabel):
                    rows_to_remove.append(row)
                    continue
            
            # 추가 확인: span이 설정된 행 감지 (확장 행의 특징)
            # 1-7번 컬럼에 아이템이 없거나 빈 텍스트인 경우 (span으로 병합된 행)
            has_real_data = False
            for col in range(1, 8):  # 1번부터 7번 컬럼까지 확인
                item = self.table_widget.item(row, col)
                if item is not None and item.text().strip():
                    has_real_data = True
                    break
            
            # 0번 컬럼에 위젯이 있지만 다른 컬럼에 실제 데이터가 없는 경우 확장 행으로 판단
            if widget is not None and not has_real_data:
                rows_to_remove.append(row)
        
        # 역순으로 제거 (인덱스 변경 방지)
        for row in reversed(rows_to_remove):
            self.table_widget.removeRow(row)

    def filter_cc_videos(self, state):
        """CC 필터 체크박스 상태에 따라 영상들을 숨기기/보이기 (확장 행 포함)"""
        is_checked = (state == 2)  # Qt.Checked == 2
        
        # 검색 결과가 없으면 체크박스를 다시 비활성화하고 리턴
        if not self.all_results:
            self.cc_filter_checkbox.setEnabled(False)
            self.cc_filter_checkbox.setChecked(False)
            return
            
        # 현재 페이지의 시작 인덱스 계산
        start_idx = (self.current_page - 1) * self.results_per_page
        
        # 현재 페이지의 데이터 행들을 정확히 찾기
        data_rows = []  # (row_index, global_idx) 형태로 저장
        current_data_count = 0  # 현재 페이지에서의 데이터 순서
        
        for row in range(self.table_widget.rowCount()):
            # 확장 행인지 확인 (cellWidget이 있으면 확장 행)
            widget = self.table_widget.cellWidget(row, 0)
            if widget:
                # 확장 행 - 건너뛰기
                continue
            
            # 원본 행인지 확인 (숨김 데이터가 있는지)
            data_item = self.table_widget.item(row, 7)
            if data_item:
                # 정확한 글로벌 인덱스 계산
                global_idx = start_idx + current_data_count
                if global_idx < len(self.all_results):
                    data_rows.append((row, global_idx))
                    current_data_count += 1
        
        # 각 원본 행에 대해 필터링 적용
        cc_count = 0
        visible_count = 0
        
        for row, global_idx in data_rows:
            result_item = self.all_results[global_idx]
            license_info = result_item.get('license', 'youtube')
            
            if license_info == 'creativeCommon':
                cc_count += 1
            
            should_hide = False
            if is_checked:
                # CC 콘텐츠만 보기: CC가 아닌 것들을 숨김
                should_hide = (license_info != 'creativeCommon')
            
            # 원본 행 숨기기/보이기
            self.table_widget.setRowHidden(row, should_hide)
            
            if not should_hide:
                visible_count += 1
            
            # 확장 행이 있다면 함께 처리
            if row + 1 < self.table_widget.rowCount():
                next_widget = self.table_widget.cellWidget(row + 1, 0)
                if next_widget:  # 다음 행이 확장 행이면
                    self.table_widget.setRowHidden(row + 1, should_hide)
        
        # 결과 라벨 업데이트 (사용자 요청 형식)
        if is_checked:
            if cc_count == 0:
                # CC 영상이 없으면 모든 행을 숨김
                for row in range(self.table_widget.rowCount()):
                    self.table_widget.setRowHidden(row, True)
                self.results_label.setText(_("결과 리스트 총 0개 (현재 페이지: 0개)"))
            else:
                self.results_label.setText(_("결과 리스트 총 {}개 (현재 페이지: {}개)").format(
                    cc_count, visible_count
                ))
        else:
            # CC 필터 해제 시 모든 행을 다시 보이게 하고 원래 상태로 복원
            for row in range(self.table_widget.rowCount()):
                self.table_widget.setRowHidden(row, False)
            self.results_label.setText(_("결과 리스트 총 {}개 (현재 페이지: {}개)").format(
                len(self.all_results), 
                min(self.results_per_page, len(self.all_results))
            ))
        

        
        # 디버깅 정보 (개발용)
        total_count = len(self.all_results) if self.all_results else 0
        print(f"🔍 CC 필터: 전체 {total_count}개 중 CC {cc_count}개 ({'ON' if is_checked else 'OFF'})")

    def _validate_api_key(self, api_key):
        """API 키 상태를 사전 검증하는 초고속 테스트 (5초 타임아웃)"""
        from urllib.parse import urlencode
        
        try:
            # 🚀 초고속 검증: HTTP GET 요청으로 직접 API 호출 (타임아웃 5초)
            # 가장 가벼운 API: search.list (quota=100이지만 빠름)
            params = {
                'part': 'snippet',
                'q': 'test',
                'type': 'video',
                'maxResults': 1,
                'key': api_key
            }
            
            url = f"https://www.googleapis.com/youtube/v3/search?{urlencode(params)}"
            
            # 5초 타임아웃으로 빠른 검증
            response = requests.get(url, timeout=5)
            
            if response.status_code == 200:
                return True, "API 키가 정상입니다."
            elif response.status_code == 403:
                # 403 오류 상세 분석
                try:
                    error_data = response.json()
                    reason = error_data.get('error', {}).get('errors', [{}])[0].get('reason', 'unknown')
                    
                    if reason == 'quotaExceeded':
                        return False, "QUOTA_EXCEEDED"
                    elif reason == 'keyInvalid' or reason == 'forbidden':
                        return False, "KEY_INVALID"
                    else:
                        return False, f"API 오류: {reason}"
                except:
                    return False, "API 접근 권한 오류"
            elif response.status_code == 400:
                return False, "KEY_INVALID"
            else:
                return False, f"HTTP 오류: {response.status_code}"
                
        except requests.exceptions.Timeout:
            return False, "연결 시간 초과 (5초) - 네트워크 상태를 확인해주세요"
        except requests.exceptions.ConnectionError:
            return False, "인터넷 연결을 확인해주세요"
        except Exception as e:
            return False, f"연결 오류: {str(e)}"

    def start_search(self):
        # --- 1단계: 모든 입력값 및 API 키 확보 ---
        api_key = self.api_key_input.text().strip()
        input_text = self.keyword_input.text().strip()
        
        # --- 2단계: 'API 사전 점검' 로직 (가장 먼저 실행) ---
        if not api_key:
            QMessageBox.warning(self, _("경고"), _("API Key를 입력해주세요."))
            return

        self.search_button.setEnabled(False)
        self.search_button.setText(_("API 키 확인 중..."))
        QApplication.processEvents() # UI 업데이트 강제

        try:
            youtube_test = build('youtube', 'v3', developerKey=api_key)
            # 가장 간단한 API 호출로 키 유효성 검사
            youtube_test.search().list(part='id', q='test', maxResults=1).execute()
        except HttpError as e:
            error_message_to_show = ""
            try:
                error_details = json.loads(e.content.decode('utf-8'))['error']
                reason = error_details.get('errors', [{}])[0].get('reason', 'unknown')
                error_message = error_details.get('message', '')
                
                if reason == 'quotaExceeded':
                    error_message_to_show = _("금일 API 할당량을 전부 소진했습니다. 다른 API Key를 이용하거나, 내일 다시 시도해주세요.")
                    QMessageBox.critical(self, _("API 할당량 초과"), error_message_to_show)
                elif reason == 'keyInvalid' or 'API key not valid' in error_message or 'badRequest' in reason:
                    error_message_to_show = _("올바르지 않은 API 키입니다. 저장된 다른 키를 이용하거나, 새로 발급받아주세요.")
                    QMessageBox.critical(self, _("API 키 오류"), error_message_to_show)
                elif e.resp.status == 400:  # HTTP 400 Bad Request
                    error_message_to_show = _("올바르지 않은 API 키입니다. 저장된 다른 키를 이용하거나, 새로 발급받아주세요.")
                    QMessageBox.critical(self, _("API 키 오류"), error_message_to_show)
                else:
                    error_message_to_show = _("올바르지 않은 API 키입니다. 저장된 다른 키를 이용하거나, 새로 발급받아주세요.")
                    QMessageBox.critical(self, _("API 키 오류"), error_message_to_show)
            except:
                # JSON 파싱 실패 시에도 API 키 오류로 간주
                error_message_to_show = _("올바르지 않은 API 키입니다. 저장된 다른 키를 이용하거나, 새로 발급받아주세요.")
                QMessageBox.critical(self, _("API 키 오류"), error_message_to_show)

            self.search_button.setEnabled(True)
            self.search_button.setText(_("검색"))
            return # 오류 발생 시, 여기서 확실하게 함수 종료
            
        # --- 3단계: 입력값 유효성 검사 ---
        if not input_text:
            QMessageBox.warning(self, _("경고"), _("검색할 키워드나 영상 URL을 입력해주세요."))
            self.search_button.setEnabled(True)
            self.search_button.setText(_("검색"))
            return
            
        # --- 4단계: (검증 완료 후) Worker 스레드 생성 및 실행 ---
        self.search_button.setText(_("검색 시작..."))
        
        # 입력값 및 필터 정리 (기존 로직)
        min_views = int(self.min_views_input.currentData())
        upload_period = self.upload_period_input.currentData()
        sort_by = self.sort_by_input.currentData()
        max_subs = self.max_subs_input.value()
        video_duration = 'any'
        if self.duration_short_radio.isChecked(): video_duration = 'short'
        elif self.duration_long_radio.isChecked(): video_duration = 'long'
        
        is_url = input_text.startswith('http')
        
        # Worker 생성 - URL과 키워드 검색 지능형 분기
        if is_url:
            # URL 유효성 검사 먼저 수행
            from urllib.parse import urlparse, parse_qs
            
            def is_valid_youtube_url(url):
                """유튜브 URL 유효성 검사"""
                try:
                    parsed = urlparse(url)
                    if parsed.netloc not in ['www.youtube.com', 'youtube.com', 'youtu.be', 'm.youtube.com']:
                        return False
                    
                    # 일반 영상 URL 패턴
                    if '/watch' in parsed.path and 'v=' in parsed.query:
                        return True
                    # 쇼츠 URL 패턴
                    if '/shorts/' in parsed.path:
                        return True
                    # 단축 URL 패턴
                    if parsed.netloc == 'youtu.be' and len(parsed.path) > 1:
                        return True
                    
                    return False
                except:
                    return False
            
            # URL 유효성 검사
            if not is_valid_youtube_url(input_text):
                QMessageBox.warning(
                    self, 
                    _("잘못된 URL"), 
                    _("올바른 유튜브 영상 URL을 입력해주세요.\n\n"
                      "지원되는 형식:\n"
                      "• https://www.youtube.com/watch?v=영상ID\n"
                      "• https://www.youtube.com/shorts/영상ID\n"
                      "• https://youtu.be/영상ID")
                )
                self.search_button.setEnabled(True)
                self.search_button.setText(_("검색"))
                return
            
            # URL 검색 확인 팝업 표시
            reply = QMessageBox.question(
                self,
                _("URL 검색 확인"),
                _("🔍 URL 검색 주의사항\n\n"
                  "• 최대 50개 영상까지만 검색됩니다\n"
                  "• 키워드 검색보다 시간이 더 오래 걸립니다 (1-3분)\n"
                  "• API 사용량이 키워드 검색보다 많습니다\n"
                  "• 검색 중 프로그램을 종료하지 마세요\n\n"
                  "계속 진행하시겠습니까?"),
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # [Task 4 Solution] MOVED UI clearing code here.
                # The UI is now cleared only AFTER the user confirms.
                self.table_widget.setRowCount(0)
                self.results_label.setText(_("결과 리스트"))
                
                # CC 필터 체크박스 비활성화 (검색 중)
                self.cc_filter_checkbox.setEnabled(False)
                self.cc_filter_checkbox.setChecked(False)
                
                self.update_pagination_ui()
            else:
                # User clicked "No", so we restore the search button and do nothing else.
                self.search_button.setEnabled(True)
                self.search_button.setText(_("검색"))
                return # Stop the search process
            
            # 진행상황 팝업 생성
            self.progress_dialog = QProgressDialog(_("URL 검색 진행 중..."), _("취소"), 0, 100, self)
            self.progress_dialog.setWindowTitle(_("검색 진행상황"))
            self.progress_dialog.setWindowModality(Qt.WindowModal)
            self.progress_dialog.setAutoClose(False)
            self.progress_dialog.setAutoReset(False)
            self.progress_dialog.setMinimumDuration(0)
            self.progress_dialog.setValue(0)
            self.progress_dialog.show()
            
            # 유사 영상 검색 - SimilarVideoWorker 사용
            self.worker = SimilarVideoWorker(
                video_url=input_text,
                upload_period=upload_period,
                min_views=min_views,
                sort_by=sort_by,
                video_duration=video_duration,
                max_subs=max_subs,
                api_key=api_key,
                translator_func=_
            )
        else: # This is a keyword search
            # Clear UI for keyword searches as before
            self.table_widget.setRowCount(0)
            self.results_label.setText(_("결과 리스트"))
            
            # CC 필터 체크박스 비활성화 (검색 중)
            self.cc_filter_checkbox.setEnabled(False)
            self.cc_filter_checkbox.setChecked(False)
            
            self.update_pagination_ui()
            
            # 키워드 검색 - 기존 Worker 사용
            self.worker = Worker(
                search_keyword=input_text,
                upload_period=upload_period,
                min_views=min_views,
                sort_by=sort_by,
                video_duration=video_duration,
                max_subs=max_subs,
                api_key=api_key,
                translator_func=_,
                similar_video_url=None,
                analyze_contribution=True
            )
        
        self.worker.finished.connect(self.on_search_finished)
        self.worker.error.connect(self.on_search_error)
        self.worker.progress.connect(self.on_search_progress)
        self.worker.start()

    def on_search_progress(self, message):
        self.search_button.setText(message)

        if hasattr(self, 'progress_dialog') and self.progress_dialog.isVisible():
            # Parse progress percentage from the message string
            import re
            match = re.search(r'\((\d+)%\)', message)
            if match:
                progress_value = int(match.group(1))
                self.progress_dialog.setValue(progress_value)

            self.progress_dialog.setLabelText(f"🔍 {message}")

            # [Task 3 Solution] Check if the user cancelled and signal the worker
            if self.progress_dialog.wasCanceled():
                if hasattr(self, 'worker') and self.worker.isRunning():
                    # Signal the worker thread to stop
                    if hasattr(self.worker, 'cancel'):
                        self.worker.cancel()

                    # Restore the UI immediately
                    self.search_button.setEnabled(True)
                    self.search_button.setText(_("검색"))
                    self.results_label.setText(_("검색이 사용자에 의해 취소되었습니다."))
                    self.progress_dialog.close()

    def on_search_finished(self, data):
        # 진행상황 팝업 닫기
        if hasattr(self, 'progress_dialog') and self.progress_dialog.isVisible():
            self.progress_dialog.setValue(100)
            self.progress_dialog.close()
        
        self.all_results = data['results']
        self.current_page = 1
        self.total_pages = math.ceil(len(self.all_results) / self.results_per_page) if self.all_results else 0
        
        # 새 검색 시 선택 상태 초기화
        self.global_selected_items.clear()
        
        # CC 필터 체크박스 초기화 및 활성화
        self.cc_filter_checkbox.setChecked(False)
        self.cc_filter_checkbox.setEnabled(len(self.all_results) > 0)  # 결과가 있을 때만 활성화
        
        self.update_table()
        self.update_pagination_ui()

        self.search_button.setEnabled(True)
        self.search_button.setText(_("검색"))
        
        if not self.all_results:
            QMessageBox.information(self, _("정보"), _("조건에 맞는 결과가 없습니다."))
        else:
            # 검색 완료 메시지와 API 한계 설명
            total_count = len(self.all_results)
            msg = _("검색 완료! 총 {}개의 결과를 찾았습니다.\n\n").format(total_count)
            
            # 유사 영상 검색인지 확인
            if hasattr(self, 'worker') and self.worker and hasattr(self.worker, 'similar_video_url') and self.worker.similar_video_url:
                msg += _("🎯 영상 URL 검색 결과:\n")
                msg += _("• 동적 하이브리드 페이지네이션으로 최대 150~200개의 유사 영상을 제공합니다.\n")
                msg += _("• 쇼츠/일반 영상을 자동 감지하여 최적의 검색 전략을 적용했습니다.\n")
                msg += _("• API 최적화를 위해 최대 200개까지의 결과만 제공합니다.\n\n")
            
            msg += _("💡 참고사항:\n")
            msg += _("• 정렬 방식에 따라 결과 개수가 4-5개 정도 차이날 수 있습니다.\n")
            msg += _("• 이는 YouTube API의 페이지네이션 특성으로 인한 것으로,\n")
            msg += _("  완전히 동일한 개수를 보장하기 어려운 한계로 인해 발생합니다.\n")
            msg += _("• 그렇기에 보다 더 정확한 트렌드 분석과 품질 높은 데이터를 제공합니다.")
            
            QMessageBox.information(self, _("검색 완료"), msg)

    def on_search_error(self, message):
        # 진행상황 팝업 닫기
        if hasattr(self, 'progress_dialog') and self.progress_dialog.isVisible():
            self.progress_dialog.close()
        
        # 구체적인 오류 코드별 맞춤형 메시지 처리
        if message == "QUOTA_EXCEEDED":
            # 할당량 초과 안내
            QMessageBox.critical(
                self, 
                _("API 할당량 초과"), 
                _("금일 API 할당량을 전부 소진했습니다.\n\n"
                  "해결 방법:\n"
                  "1. 내일까지 기다리기 (GMT 기준으로 할당량이 리셋됩니다)\n"
                  "2. 다른 API 키 사용하기\n"
                  "3. Google Cloud Console에서 할당량 증가 요청하기")
            )
        elif message == "KEY_INVALID":
            # 잘못된 키 안내
            QMessageBox.critical(
                self, 
                _("API 키 오류"), 
                _("올바르지 않은 API 키입니다.\n\n"
                  "해결 방법:\n"
                  "1. Google Cloud Console에서 새 API 키 생성\n"
                  "2. YouTube Data API v3가 활성화되어 있는지 확인\n"
                  "3. API 키에 적절한 권한이 설정되어 있는지 확인")
            )
        else:
            # 일반 오류 메시지
            QMessageBox.critical(self, _("오류"), message)
        
        self.search_button.setEnabled(True)
        self.search_button.setText(_("검색"))
        self.results_label.setText(_("결과 리스트"))
        self.update_pagination_ui()

    def update_table(self):
        if not self.all_results:
            self.table_widget.setRowCount(0)
            self.results_label.setText(_("결과 리스트"))
            return

        # 페이지 변경시 확장된 행 완전히 제거
        self._remove_all_expanded_rows()
        self.expanded_videos.clear()
        
        # 페이지 변경 시 마지막 선택된 행 초기화
        self.last_selected_row = -1

        # 현재 페이지에 해당하는 결과 계산
        start_idx = (self.current_page - 1) * self.results_per_page
        end_idx = min(start_idx + self.results_per_page, len(self.all_results))
        current_page_results = self.all_results[start_idx:end_idx]

        self.table_widget.setRowCount(len(current_page_results))
        
        # 행 헤더에 전체 순서 번호 설정
        vertical_headers = []
        for row, item in enumerate(current_page_results):
            # 전체 순서 번호 계산 (1부터 시작)
            global_row_number = (self.current_page - 1) * self.results_per_page + row + 1
            vertical_headers.append(str(global_row_number))
            
            # 체크박스 - 여백 없음
            checkbox = QCheckBox()
            
            # Shift 키 범위 선택 기능 추가
            checkbox.clicked.connect(lambda checked, r=row: self._on_checkbox_clicked(r, checked))
            
            cell_widget = QWidget()
            layout = QHBoxLayout(cell_widget)
            layout.addWidget(checkbox)
            layout.setAlignment(Qt.AlignCenter)  # type: ignore
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setSpacing(0)
            self.table_widget.setCellWidget(row, 0, cell_widget)
            
            # 데이터 입력
            self.table_widget.setItem(row, 1, QTableWidgetItem(item['published_at']))
            self.table_widget.setItem(row, 2, QTableWidgetItem(f"{item['view_count']:,}"))

            # 분량 - 중앙 정렬
            duration_item = QTableWidgetItem(item['duration_formatted'])
            duration_item.setTextAlignment(Qt.AlignCenter)  # type: ignore
            self.table_widget.setItem(row, 3, duration_item)

            self.table_widget.setItem(row, 4, QTableWidgetItem(item['title']))
            self.table_widget.setItem(row, 5, QTableWidgetItem(item['channel_title']))
            self.table_widget.setItem(row, 6, QTableWidgetItem(f"{item['subscriber_count']:,}"))
            
            # 숨김 데이터 (video_id, channel_id 저장)
            data_item = QTableWidgetItem(f"{item['video_id']}|{item['channel_id']}")
            self.table_widget.setItem(row, 7, data_item)
        
        # 행 헤더 설정
        self.table_widget.setVerticalHeaderLabels(vertical_headers)

        # 결과 라벨 업데이트
        total_count = len(self.all_results)
        showing_count = len(current_page_results)
        self.results_label.setText(_("결과 리스트 총 {}개 (현재 페이지: {}개)").format(total_count, showing_count))
        
        # 페이지 로드 후 선택 상태 복원
        self._update_current_page_checkboxes()
        self._update_selected_count_label()

    def _on_checkbox_clicked(self, row, checked):
        """체크박스 클릭 이벤트 핸들러 (Shift 키 범위 선택 지원)"""
        from PySide6.QtWidgets import QApplication
        
        # Shift 키가 눌려있고, 이전에 선택된 행이 있는 경우 범위 선택
        if (QApplication.keyboardModifiers() & Qt.KeyboardModifier.ShiftModifier and 
            self.last_selected_row != -1 and 
            self.last_selected_row != row):
            
            # 범위 계산
            start_row = min(self.last_selected_row, row)
            end_row = max(self.last_selected_row, row)
            
            # 범위 내의 모든 체크박스를 현재 체크박스와 동일한 상태로 설정
            for r in range(start_row, end_row + 1):
                if r < self.table_widget.rowCount():
                    cell_widget = self.table_widget.cellWidget(r, 0)
                    if cell_widget:
                        checkbox = cell_widget.findChild(QCheckBox)
                        if checkbox:
                            checkbox.setChecked(checked)
        
        # 현재 행을 마지막 선택된 행으로 저장
        self.last_selected_row = row
        
        # 전역 선택 상태 업데이트
        self._update_global_selection_from_current_page()
        self._update_selected_count_label()
    
    def _update_global_selection_from_current_page(self):
        """현재 페이지의 체크박스 상태를 전역 선택 상태에 반영"""
        start_idx = (self.current_page - 1) * self.results_per_page
        
        for row in range(self.table_widget.rowCount()):
            global_idx = start_idx + row
            if global_idx < len(self.all_results):
                cell_widget = self.table_widget.cellWidget(row, 0)
                if cell_widget:
                    checkbox = cell_widget.findChild(QCheckBox)
                    if checkbox:
                        if checkbox.isChecked():
                            self.global_selected_items.add(global_idx)
                        else:
                            self.global_selected_items.discard(global_idx)
    
    def _update_current_page_checkboxes(self):
        """전역 선택 상태를 현재 페이지의 체크박스에 반영"""
        start_idx = (self.current_page - 1) * self.results_per_page
        
        for row in range(self.table_widget.rowCount()):
            global_idx = start_idx + row
            if global_idx < len(self.all_results):
                cell_widget = self.table_widget.cellWidget(row, 0)
                if cell_widget:
                    checkbox = cell_widget.findChild(QCheckBox)
                    if checkbox:
                        checkbox.setChecked(global_idx in self.global_selected_items)
    
    def _update_selected_count_label(self):
        """선택된 개수를 라벨에 표시"""
        count = len(self.global_selected_items)
        self.selected_count_label.setText(_("선택됨: {}개").format(count))

    def update_pagination_ui(self):
        if self.total_pages <= 1:
            self.prev_page_button.setEnabled(False)
            self.next_page_button.setEnabled(False)
            self.page_info_label.setText("")
        else:
            self.prev_page_button.setEnabled(self.current_page > 1)
            self.next_page_button.setEnabled(self.current_page < self.total_pages)
            self.page_info_label.setText(f"{self.current_page}/{self.total_pages}")

    def prev_page(self):
        if self.current_page > 1:
            # 현재 페이지의 선택 상태 저장
            self._update_global_selection_from_current_page()
            self.current_page -= 1
            self.update_table()
            self.update_pagination_ui()

    def next_page(self):
        if self.current_page < self.total_pages:
            # 현재 페이지의 선택 상태 저장
            self._update_global_selection_from_current_page()
            self.current_page += 1
            self.update_table()
            self.update_pagination_ui()

    def get_selected_rows(self) -> List[int]:
        """현재 페이지에서 선택된 행들을 반환 (페이지 내 인덱스)"""
        selected_rows = []
        start_idx = (self.current_page - 1) * self.results_per_page
        
        for row in range(self.table_widget.rowCount()):
            global_idx = start_idx + row
            if global_idx in self.global_selected_items:
                selected_rows.append(row)
        return selected_rows
    
    def get_all_selected_global_indices(self) -> List[int]:
        """전체 결과에서 선택된 모든 항목의 전역 인덱스를 반환"""
        return sorted(list(self.global_selected_items))
    
    def toggle_select_all(self, state: bool):
        if state:
            # 모두 선택: 전체 결과의 모든 인덱스를 선택
            self.global_selected_items = set(range(len(self.all_results)))
        else:
            # 모두 해제: 전체 선택 해제
            self.global_selected_items.clear()
        
        # 현재 페이지의 체크박스 상태 업데이트
        self._update_current_page_checkboxes()
        self._update_selected_count_label()

    def export_to_excel(self):
        if self.table_widget.rowCount() == 0:
            QMessageBox.warning(self, _("경고"), _("추출할 데이터가 없습니다."))
            return

        # 전역 선택된 항목 확인
        selected_global_indices = self.get_all_selected_global_indices()
        if not selected_global_indices:
            QMessageBox.warning(self, _("경고"), _("추출할 항목을 선택해주세요."))
            return

        # 설정된 경로가 있으면 사용, 없으면 다운로드 폴더 사용
        save_dir = self.excel_save_path or self.get_downloads_folder()
        os.makedirs(save_dir, exist_ok=True)

        # 헤더 생성 (기여도 관련 컬럼 추가)
        headers = [_('썸네일'), _('업로드 날짜'), _('조회수'), _('기여도(성과)'), _('채널 평균 조회수'), _('분량'), _('좋아요'), _('댓글'), _('제목'), _('채널명'), _('구독자 수'), _('영상 링크'), _('채널 링크')]
        
        # 선택된 전역 인덱스에 해당하는 데이터 수집
        selected_data = []
        for global_index in selected_global_indices:
            if global_index < len(self.all_results):
                item = self.all_results[global_index]
                
                # 기여도 데이터 준비 (견고한 처리)
                contribution_text = "N/A"
                channel_avg_text = "N/A"
                
                # 기여도 데이터가 존재하고 유효한 경우만 표시
                try:
                    if ('contribution_score' in item and 'channel_avg_views' in item and 
                        item['contribution_score'] is not None and item['channel_avg_views'] is not None and
                        item['contribution_score'] > 0 and item['channel_avg_views'] > 0):
                        contribution_text = f"{item['contribution_score']:.1f}%"
                        channel_avg_text = f"{item['channel_avg_views']:,}"
                except (KeyError, TypeError, ValueError):
                    # 데이터 오류 시 N/A 유지
                    pass
                
                row_data = [
                    '',  # 썸네일 자리 (이미지는 별도로 삽입)
                    item.get('published_at', ''),
                    f"{item.get('view_count', 0):,}",
                    contribution_text,      # 기여도(성과) - 항상 포함
                    channel_avg_text,       # 채널 평균 조회수 - 항상 포함
                    item.get('duration_formatted', ''),
                    f"{item.get('like_count', 0):,}",
                    f"{item.get('comment_count', 0):,}",
                    item.get('title', ''),
                    item.get('channel_title', ''),
                    f"{item.get('subscriber_count', 0):,}",
                    f"https://www.youtube.com/watch?v={item.get('video_id', '')}",
                    f"https://www.youtube.com/channel/{item.get('channel_id', '')}"
                ]
                selected_data.append((row_data, item))

        filename = os.path.join(save_dir, f"youtube_search_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        # 진행상황 다이얼로그 생성 (썸네일 추출과 동일한 방식)
        progress = QProgressDialog(_("데이터를 추출중입니다. 잠시만 기다려주십시오."), _("취소"), 0, len(selected_data), self)
        progress.setModal(True)
        progress.setAutoClose(False)
        progress.setAutoReset(False)
        progress.show()
        
        try:
            data = [item[0] for item in selected_data]  # 행 데이터만 추출
            items = [item[1] for item in selected_data]  # 원본 아이템들
            
            # 진행상황 업데이트를 위한 콜백 함수
            def update_progress(current, total):
                if progress.wasCanceled():
                    return False
                progress.setValue(current)
                progress.setLabelText(_("데이터를 추출중입니다. 잠시만 기다려주십시오. ({}/{})").format(current, total))
                QApplication.processEvents()  # UI 업데이트
                return True
            
            self._create_excel_with_thumbnails(data, headers, filename, items, update_progress)
            
            # 진행 다이얼로그 닫기
            progress.close()
            
            # 완료 메시지 표시
            QMessageBox.information(
                self, _("성공"), 
                _("엑셀 추출이 완료되었습니다!\n'{}' 파일로 저장되었습니다.\n총 {}개 결과가 저장되었습니다.").format(os.path.basename(filename), len(data))
            )
            
        except Exception as e:
            # 진행 다이얼로그 닫기
            progress.close()
            QMessageBox.critical(self, _("엑셀 저장 오류"), _("파일 저장 중 오류가 발생했습니다: {e}").format(e))
    

    
    def _create_excel_with_thumbnails(self, data, headers, filename, items=None, progress_callback=None):
        """썸네일을 포함한 엑셀 파일 생성"""
        if not PIL_AVAILABLE:
            # PIL이 없으면 기본 방식으로 저장
            df = pd.DataFrame(data, columns=headers)
            df.to_excel(filename, index=False)
            return
        
        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "YouTube"
            
            # 헤더 삽입
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                if cell is not None:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 데이터 삽입
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if cell is not None:
                        # 제목(I열, 9번째)과 채널명(J열, 10번째)은 좌측 정렬, 나머지는 중앙 정렬
                        if col_idx in [9, 10]:  # 제목, 채널명
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 열 너비 설정 (새로운 컬럼들 반영)
            ws.column_dimensions['A'].width = 18  # 썸네일 (너비 확장)
            ws.column_dimensions['B'].width = 12  # 업로드 날짜
            ws.column_dimensions['C'].width = 12  # 조회수
            ws.column_dimensions['D'].width = 12  # 기여도(성과)
            ws.column_dimensions['E'].width = 15  # 채널 평균 조회수
            ws.column_dimensions['F'].width = 8   # 분량
            ws.column_dimensions['G'].width = 10  # 좋아요
            ws.column_dimensions['H'].width = 8   # 댓글
            ws.column_dimensions['I'].width = 50  # 제목 (넓게)
            ws.column_dimensions['J'].width = 20  # 채널명
            ws.column_dimensions['K'].width = 12  # 구독자 수
            ws.column_dimensions['L'].width = 40  # 영상 링크
            ws.column_dimensions['M'].width = 40  # 채널 링크
            
            # 썸네일 이미지 다운로드 및 삽입 (병렬 처리)
            if items:
                self._download_thumbnails_parallel(ws, items, progress_callback)
            else:
                # items가 없으면 self.all_results에서 해당하는 항목들 찾기
                self._download_thumbnails_parallel(ws, None, progress_callback)
        
        # 파일 저장
        wb.save(filename)
    
    def _download_thumbnails_parallel(self, ws, items=None, progress_callback=None):
        """병렬로 썸네일 다운로드 및 삽입"""
        # items가 주어지지 않으면 self.all_results 사용
        if items is None:
            items = self.all_results
        def download_and_insert_thumbnail(args):
            idx, item = args
            try:
                thumbnail_url = item.get('thumbnail_url', '')
                if not thumbnail_url:
                    return None
                
                # 썸네일 이미지 다운로드 (더 빠른 처리를 위한 최적화)
                response = requests.get(thumbnail_url, timeout=8, stream=True)
                if response.status_code == 200:
                    # PIL Image로 변환
                    img = Image.open(io.BytesIO(response.content))
                    
                    # 이미지 크기 조정 - 컬럼 너비에 맞게 조정 (18 * 7 = 126px 정도)
                    img.thumbnail((126, 95), Image.Resampling.NEAREST)
                    
                    # RGB 모드로 변환 (PNG 최적화)
                    if img.mode != 'RGB':
                        img = img.convert('RGB')
                    
                    # 임시 파일로 저장 (JPEG로 저장하여 용량 최적화)
                    img_io = io.BytesIO()
                    img.save(img_io, format='JPEG', quality=85, optimize=True)
                    img_io.seek(0)
                    
                    # openpyxl 이미지 객체 생성 - 컬럼 너비에 맞게 크기 설정
                    excel_img = OpenpyxlImage(img_io)
                    excel_img.width = 126  # 썸네일 컬럼 너비(18)에 맞춤
                    excel_img.height = 95   # 비율에 맞춰 조정
                    
                    return idx, excel_img
                    
            except Exception as e:
                return None
        
        # 병렬로 썸네일 다운로드
        completed_count = 0
        with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
            thumbnail_args = [(idx, item) for idx, item in enumerate(items)]
            future_to_idx = {executor.submit(download_and_insert_thumbnail, args): args[0] for args in thumbnail_args}
            
            for future in concurrent.futures.as_completed(future_to_idx):
                try:
                    result = future.result(timeout=15)  # 15초 타임아웃
                    if result is not None:
                        idx, excel_img = result
                        
                        # 행 높이 설정 (이미지에 맞춤)
                        ws.row_dimensions[idx + 2].height = 75  # 이미지 높이에 맞춤
                        
                        # 이미지를 엑셀에 삽입 (A열, 해당 행)
                        cell = f'A{idx + 2}'
                        ws.add_image(excel_img, cell)
                    
                    # 진행상황 업데이트
                    completed_count += 1
                    if progress_callback:
                        if not progress_callback(completed_count, len(items)):
                            return  # 취소된 경우 중단
                        
                except Exception as e:
                    completed_count += 1
                    if progress_callback:
                        if not progress_callback(completed_count, len(items)):
                            return  # 취소된 경우 중단
                    continue

    def open_channel_link(self):
        selected_global_indices = self.get_all_selected_global_indices()
        if not selected_global_indices:
            QMessageBox.warning(self, _("경고"), _("채널을 선택해주세요."))
            return
        
        for global_index in selected_global_indices:
            if global_index < len(self.all_results):
                item = self.all_results[global_index]
                channel_id = item['channel_id']
                webbrowser.open_new_tab(f"https://www.youtube.com/channel/{channel_id}")

    def open_video_link(self):
        selected_global_indices = self.get_all_selected_global_indices()
        if not selected_global_indices:
            QMessageBox.warning(self, _("경고"), _("영상을 선택해주세요."))
            return

        for global_index in selected_global_indices:
            if global_index < len(self.all_results):
                item = self.all_results[global_index]
                video_id = item['video_id']
                webbrowser.open_new_tab(f"https://www.youtube.com/watch?v={video_id}")

    def download_thumbnails(self):
        selected_global_indices = self.get_all_selected_global_indices()
        if not selected_global_indices:
            QMessageBox.warning(self, _("경고"), _("다운로드할 영상을 선택해주세요."))
            return
        
        # 설정된 경로가 있으면 사용, 없으면 다운로드 폴더 사용
        download_dir = self.thumbnail_save_path or self.get_downloads_folder()
        os.makedirs(download_dir, exist_ok=True)
        
        # 진행상황 다이얼로그 생성
        progress = QProgressDialog(_("데이터를 추출중입니다. 잠시만 기다려주십시오."), _("취소"), 0, len(selected_global_indices), self)
        progress.setModal(True)
        progress.setAutoClose(False)
        progress.setAutoReset(False)
        progress.show()
        
        success_count = 0
        for idx, global_index in enumerate(selected_global_indices):
            # 진행상황 업데이트
            if progress.wasCanceled():
                break
            progress.setValue(idx)
            progress.setLabelText(_("데이터를 추출중입니다. 잠시만 기다려주십시오. ({}/{})").format(idx + 1, len(selected_global_indices)))
            QApplication.processEvents()  # UI 업데이트
            try:
                if global_index < len(self.all_results):
                    item = self.all_results[global_index]
                    video_id = item['video_id']
                    title = item['title']
                    
                    # 안전한 파일명 생성
                    safe_title = "".join(c for c in title if c.isalnum() or c in "._- ").strip()
                    safe_title = safe_title[:50]  # 파일명 길이 제한
                    
                    # 썸네일 URL 시도 (고화질 -> 표준 -> 기본)
                    urls = [
                        f"https://img.youtube.com/vi/{video_id}/maxresdefault.jpg",
                        f"https://img.youtube.com/vi/{video_id}/hqdefault.jpg",
                        f"https://img.youtube.com/vi/{video_id}/default.jpg"
                    ]
                    
                    for url in urls:
                        response = requests.get(url, timeout=10)
                        if response.status_code == 200 and len(response.content) > 1000:  # 최소 크기 확인
                            filepath = os.path.join(download_dir, f"{safe_title}_{video_id}.jpg")
                            with open(filepath, 'wb') as f:
                                f.write(response.content)
                            success_count += 1
                            break
            except Exception as e:
                print(f"썸네일 다운로드 오류: {e}")
                continue
        
        # 진행 다이얼로그 닫기
        progress.close()
        
        if success_count > 0:
            QMessageBox.information(
                self, _("성공"), 
                _("썸네일 추출이 완료되었습니다!\n선택한 썸네일 {}개를 '{}' 폴더에 저장했습니다.").format(success_count, os.path.basename(download_dir))
            )
        else:
            QMessageBox.warning(self, _("실패"), _("썸네일 다운로드에 실패했습니다."))
    
    # === 경로 설정 관련 메서드들 ===
    def get_downloads_folder(self) -> str:
        """다운로드 폴더 경로를 반환합니다."""
        return os.path.join(os.path.expanduser("~"), "Downloads")
    
    def load_settings(self):
        """설정 파일에서 경로 설정과 마지막 사용 API Key를 로드합니다."""
        try:
            if os.path.exists(path_manager.SETTINGS_PATH):
                with open(path_manager.SETTINGS_PATH, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    self.excel_save_path = settings.get("excel_save_path")
                    self.thumbnail_save_path = settings.get("thumbnail_save_path")
                    self.last_used_api_key = settings.get("last_used_api_key", "")
            else:
                self.excel_save_path = None
                self.thumbnail_save_path = None
                self.last_used_api_key = ""
        except Exception as e:
            print(f"설정 로드 오류: {e}")
            self.excel_save_path = None
            self.thumbnail_save_path = None
            self.last_used_api_key = ""
    
    def save_settings(self):
        """현재 경로 설정과 마지막 사용 API Key를 파일에 저장합니다."""
        settings = {
            "excel_save_path": self.excel_save_path,
            "thumbnail_save_path": self.thumbnail_save_path,
            "last_used_api_key": self.last_used_api_key
        }
        try:
            with open(path_manager.SETTINGS_PATH, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"설정 저장 오류: {e}")
    
    def select_excel_path(self):
        """엑셀 파일 저장 경로를 선택합니다."""
        folder = QFileDialog.getExistingDirectory(
            self, 
            _("엑셀 파일 저장 경로 선택"),
            self.excel_save_path or self.get_downloads_folder()
        )
        if folder:
            self.excel_save_path = folder
            self.save_settings()
            self.update_path_display()
    
    def select_thumbnail_path(self):
        """썸네일 저장 경로를 선택합니다."""
        folder = QFileDialog.getExistingDirectory(
            self, 
            _("썸네일 저장 경로 선택"),
            self.thumbnail_save_path or self.get_downloads_folder()
        )
        if folder:
            self.thumbnail_save_path = folder
            self.save_settings()
            self.update_path_display()
    
    def update_path_display(self):
        """경로 표시 라벨을 업데이트합니다."""
        # 엑셀 경로 표시
        if self.excel_save_path:
            self.excel_path_label.setText(_("엑셀: {}").format(self.excel_save_path))
            self.excel_path_label.setStyleSheet("color: #0078D4; font-size: 9px;")
        else:
            self.excel_path_label.setText(_("엑셀: 다운로드 폴더"))
            self.excel_path_label.setStyleSheet("color: #666666; font-size: 9px;")
        
        # 썸네일 경로 표시
        if self.thumbnail_save_path:
            self.thumbnail_path_label.setText(_("썸네일: {}").format(self.thumbnail_save_path))
            self.thumbnail_path_label.setStyleSheet("color: #0078D4; font-size: 9px;")
        else:
            self.thumbnail_path_label.setText(_("썸네일: 다운로드 폴더"))
            self.thumbnail_path_label.setStyleSheet("color: #666666; font-size: 9px;")

    def check_first_run(self):
        """최초 실행 확인 및 언어 설정"""
        config_file = path_manager.CONFIG_PATH
        if not os.path.exists(config_file):
            # 최초 실행 - 영어로 고정된 언어 선택 팝업 표시
            self.show_first_run_language_dialog()
        else:
            # 기존 설정 로드
            self.load_language_from_config()
    
    def show_first_run_language_dialog(self):
        """최초 실행 시 영어로 고정된 언어 선택 다이얼로그"""
        from PySide6.QtWidgets import QVBoxLayout, QHBoxLayout, QListWidget, QListWidgetItem
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Language Selection")  # 영어로 고정
        dialog.setModal(True)
        # 반응형 다이얼로그 설정 (기본 크기 400x300, 최소 크기 동일)
        dialog.resize(400, 300)
        dialog.setMinimumSize(400, 300)
        
        layout = QVBoxLayout()
        
        # 안내 문구 (영어로 고정)
        info_label = QLabel("Please select your language:")
        info_label.setStyleSheet("font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(info_label)
        
        # 언어 목록
        self.language_list = QListWidget()
        self.language_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        
        # 지원 언어 목록 (로컬 변수로 정의)
        supported_languages = {
            'ko': '한국어',
            'en': 'English',
            'ja': '日본語',
            'zh': '中文',
            'es': 'Español',
            'fr': 'Français',
            'de': 'Deutsch',
            'ru': 'Русский',
            'pt': 'Português',
            'it': 'Italiano',
            'hi': 'हिन्दी (Hindi)',
            'ar': 'العربية (Arabic)',
            'tr': 'Türkçe',
            'th': 'ไทย (Thai)',
            'id': 'Bahasa Indonesia',
            'vi': 'Tiếng Việt',
            'nl': 'Nederlands',
            'pl': 'Polski',
            'sv': 'Svenska',
            'da': 'Dansk',
            'no': 'Norsk',
            'fi': 'Suomi',
            'uk': 'Українська'
        }
        
        # 언어 항목 추가
        for code, name in supported_languages.items():
            item = QListWidgetItem(name)
            item.setData(Qt.ItemDataRole.UserRole, code)
            self.language_list.addItem(item)
        
        # 기본으로 한국어 선택
        for i in range(self.language_list.count()):
            item = self.language_list.item(i)
            if item.data(Qt.ItemDataRole.UserRole) == 'ko':
                self.language_list.setCurrentItem(item)
                break
        
        layout.addWidget(self.language_list)
        
        # 버튼
        button_layout = QHBoxLayout()
        ok_button = QPushButton("OK")  # 영어로 고정
        ok_button.setStyleSheet("font-weight: bold;")
        
        cancel_button = QPushButton("Cancel")  # 영어로 고정
        
        button_layout.addStretch()
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        dialog.setLayout(layout)
        
        def on_ok():
            current_item = self.language_list.currentItem()
            if current_item:
                selected_language = current_item.data(Qt.ItemDataRole.UserRole)
                self.current_language = selected_language
                self.save_language_to_config()
                setup_i18n(selected_language)
                dialog.accept()
            else:
                QMessageBox.warning(dialog, "Warning", "Please select a language.")  # 영어로 고정
        
        def on_cancel():
            # 취소 시 기본 한국어 설정
            self.current_language = 'ko'
            self.save_language_to_config()
            setup_i18n('ko')
            dialog.reject()
        
        ok_button.clicked.connect(on_ok)
        cancel_button.clicked.connect(on_cancel)
        
        dialog.exec()
    
    def load_language_from_config(self):
        """config.json에서 언어 설정을 로드합니다."""
        try:
            with open(path_manager.CONFIG_PATH, 'r', encoding='utf-8') as f:
                config = json.load(f)
                self.current_language = config.get('language', 'ko')
                setup_i18n(self.current_language)
        except Exception as e:
            print(f"언어 설정 로드 오류: {e}")
            self.current_language = 'ko'
            setup_i18n('ko')
    
    def save_language_to_config(self):
        """언어 설정을 config.json에 저장합니다."""
        try:
            config = {}
            if os.path.exists(path_manager.CONFIG_PATH):
                with open(path_manager.CONFIG_PATH, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            
            config['language'] = self.current_language
            
            with open(path_manager.CONFIG_PATH, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"언어 설정 저장 오류: {e}")
            
    def open_language_manager(self):
        """언어 관리자를 엽니다."""
        dialog = LanguageManager(self)
        if dialog.exec():
            selected_language = dialog.get_selected_language()
            if selected_language != self.current_language:
                self.current_language = selected_language
                self.save_language_to_config()  # config.json에 저장
                self.update_language_button()
                QMessageBox.information(self, _("언어 변경"), _("재시작 시 적용됩니다."))
                
    def apply_translations(self):
        """번역을 적용합니다."""
        # 현재는 기본 구현만 제공
        # 실제 번역 파일이 있을 때 QTranslator.load() 사용
        pass
        
    def update_language_button(self):
        """언어 버튼 텍스트를 업데이트합니다."""
        if hasattr(self, 'language_button'):
            self.language_button.setText("Language")

class FirstRunLanguageDialog(QDialog):
    """처음 실행시 전용 언어 설정 다이얼로그"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Language Selection")  # 영어로 고정
        self.setModal(True)
        # 반응형 다이얼로그 설정 (기본 크기 400x280, 최소 크기 동일)
        self.resize(400, 280)
        self.setMinimumSize(400, 280)
        
        # 지원 언어 목록
        self.supported_languages = {
            'ko': '한국어',
            'en': 'English',
            'ja': '日本語',
            'zh': '中文',
            'es': 'Español',
            'fr': 'Français',
            'de': 'Deutsch',
            'ru': 'Русский',
            'pt': 'Português',
            'it': 'Italiano',
            'hi': 'हिन्दी (Hindi)',
            'ar': 'العربية (Arabic)',
            'tr': 'Türkçe',
            'th': 'ไทย (Thai)',
            'id': 'Bahasa Indonesia',
            'vi': 'Tiếng Việt',
            'nl': 'Nederlands',
            'pl': 'Polski',
            'sv': 'Svenska',
            'da': 'Dansk',
            'no': 'Norsk',
            'fi': 'Suomi',
            'uk': 'Українська'
        }
        
        self.selected_language = 'ko'  # 기본 언어
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # 안내 문구 (영어로 고정)
        info_label = QLabel("Please select your language:")
        info_label.setStyleSheet("font-weight: bold; margin-bottom: 15px; font-size: 14px;")
        layout.addWidget(info_label)
        
        # 언어 목록
        self.language_list = QListWidget()
        self.language_list.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        
        # 언어 항목 추가
        for code, name in self.supported_languages.items():
            item = QListWidgetItem(name)
            item.setData(Qt.ItemDataRole.UserRole, code)
            self.language_list.addItem(item)
        
        # 기본으로 한국어 선택
        for i in range(self.language_list.count()):
            item = self.language_list.item(i)
            if item.data(Qt.ItemDataRole.UserRole) == 'ko':
                self.language_list.setCurrentItem(item)
                break
        
        layout.addWidget(self.language_list)
        
        # OK 버튼만 있음 (Cancel 버튼 없음)
        button_layout = QHBoxLayout()
        
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept_language)
        self.ok_button.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 30px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #106EBE;
            }
            QPushButton:pressed {
                background-color: #005A9E;
            }
        """)
        
        button_layout.addStretch()
        button_layout.addWidget(self.ok_button)
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
    def accept_language(self):
        """언어 선택을 확인합니다."""
        current_item = self.language_list.currentItem()
        if current_item:
            self.selected_language = current_item.data(Qt.ItemDataRole.UserRole)
            
            # config.json에 즉시 저장
            config = {'language': self.selected_language}
            with open(path_manager.CONFIG_PATH, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            # 언어 즉시 적용
            setup_i18n(self.selected_language)
            
            self.accept()
        else:
            QMessageBox.warning(self, "Warning", "Please select a language.")
            
    def get_selected_language(self):
        """선택된 언어 코드를 반환합니다."""
        return self.selected_language

if __name__ == '__main__':
    # ❗️❗️❗️ 이 부분이 모든 문제를 해결하는 핵심 코드입니다 ❗️❗️❗️
    # .exe의 실제 위치를 찾아서 작업 경로로 설정합니다.
    if getattr(sys, 'frozen', False):
        application_path = os.path.dirname(sys.executable)
    else:
        application_path = os.path.dirname(os.path.abspath(__file__))
    os.chdir(application_path)
    # ❗️❗️❗️ 여기까지가 추가된 부분입니다 ❗️❗️❗️

    app = QApplication(sys.argv)
    
    # === 라이선스 인증 시스템 (최종 완성본) ===
    import license_manager
    # ... (이하 코드는 그대로 유지) ...

    # 1. 라이선스 시스템 초기화
    if license_manager.initialize_license_system():
        # -> 라이선스가 유효하거나, 방금 새로 인증에 성공한 상태

        # 2. 새로 인증되었는지 확인 (첫 실행 언어 설정용)
        if license_manager.is_license_just_activated():
            print("새 라이선스 활성화 감지. 언어 설정 창 표시...")
            dialog = FirstRunLanguageDialog(None)
            if not dialog.exec():  # ❗️'X'를 누르면 False가 반환됩니다.
                # 사용자가 창을 닫으면 프로그램 종료
                print("언어 설정이 취소되어 프로그램을 종료합니다.")
                sys.exit(0)
            # 'OK'를 눌렀을 때만 아래 코드 실행
            license_manager.reset_license_activation_flag()

        # 3. 언어 설정 로드 및 적용
        try:
            with open(path_manager.CONFIG_PATH, 'r', encoding='utf-8') as f:
                config = json.load(f)
                language = config.get('language', 'ko')
        except (FileNotFoundError, json.JSONDecodeError):
            language = 'ko'
        
        setup_i18n(language)

        # 4. 메인 프로그램 실행
        print("메인 윈도우 생성 중...")
        main_win = MainWindow()
        
        print("메인 윈도우 표시...")
        main_win.show()
        print("프로그램 시작 완료!")
        sys.exit(app.exec())
    else:
        # 라이선스가 유효하지 않거나 체험판이 만료되면 프로그램 종료
        print("라이선스가 유효하지 않거나 체험판이 만료되어 프로그램을 종료합니다.")
        sys.exit(0)